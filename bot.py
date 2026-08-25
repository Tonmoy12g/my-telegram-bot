from keep_alive import keep_alive
keep_alive()


# -*- coding: utf-8 -*-
"""
Trust Vault Mails Style - Telegram Micro Tasking Bot (Ultimate Edition)
Fully Optimized, High-Speed Broadcast, Auto-Delete Credentials, Multi-Channel Force Join, 
Dynamic Withdraw Control (ON/OFF & Min/Max), Instant Referral Alerts & Enhanced UI.
"""

import asyncio
import datetime
import html
import io
import logging
import os
import re
import sqlite3
import sys
import secrets
import random
import warnings
import time
from typing import Dict, List, Optional, Tuple, Union
from contextlib import contextmanager
from queue import Queue
import threading

import pyotp

# Try importing openpyxl for Excel spreadsheet generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from telegram import (
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.constants import ParseMode
from telegram.warnings import PTBUserWarning
from telegram.error import Forbidden, BadRequest, RetryAfter, TelegramError
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

warnings.filterwarnings("ignore", category=PTBUserWarning)

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

# -----------------------------------------------------------------------------
# LOGGING SETUP
# -----------------------------------------------------------------------------
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("bot.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("Nexora_Work_Style")

# -----------------------------------------------------------------------------
# DATABASE CONNECTION POOL & INIT
# -----------------------------------------------------------------------------
DB_FILE = "bot_database.db"

class DatabasePool:
    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self._pool = Queue()
        self._max_connections = 50
        self._min_connections = 5
        self._create_initial_pool()
        logger.info(f"Database pool initialized with {self._min_connections} connections")

    def _create_initial_pool(self):
        for _ in range(self._min_connections):
            conn = self._create_connection()
            self._pool.put(conn)

    def _create_connection(self):
        conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=30.0)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size=10000")
        conn.execute("PRAGMA temp_store=MEMORY")
        return conn

    @contextmanager
    def get_connection(self):
        conn = None
        try:
            conn = self._pool.get(timeout=10.0)
            yield conn
        except Exception as e:
            if conn:
                self._pool.put(conn)
            raise e
        finally:
            if conn:
                self._pool.put(conn)

_db_pool = DatabasePool()

def get_db():
    return _db_pool.get_connection()

def get_db_direct():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def init_db():
    with get_db_direct() as conn:
        cursor = conn.cursor()

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            balance REAL DEFAULT 0.0,
            referral_code TEXT UNIQUE,
            referred_by INTEGER,
            ref_count INTEGER DEFAULT 0,
            ref_earnings REAL DEFAULT 0.0,
            is_banned INTEGER DEFAULT 0,
            joined_at TIMESTAMP,
            language TEXT DEFAULT 'bn'
        )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_users_referral_code ON users(referral_code)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_users_referred_by ON users(referred_by)")

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS task_submissions (
            submission_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            task_type TEXT,
            submitted_username TEXT,
            submitted_password TEXT,
            secret_key_2fa TEXT,
            cookies_data TEXT DEFAULT '',
            status TEXT DEFAULT 'Pending',
            reward_amount REAL DEFAULT 0.0,
            reject_reason TEXT DEFAULT '',
            is_archived INTEGER DEFAULT 0,
            created_at TIMESTAMP,
            updated_at TIMESTAMP
        )
        """)
        
        try:
            cursor.execute("ALTER TABLE task_submissions ADD COLUMN is_archived INTEGER DEFAULT 0")
        except sqlite3.OperationalError:
            pass

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_submissions_user_id ON task_submissions(user_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_submissions_status ON task_submissions(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_submissions_username ON task_submissions(submitted_username)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_submissions_archived ON task_submissions(is_archived)")

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS admin_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS withdrawals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            method TEXT,
            number TEXT,
            amount REAL,
            status TEXT DEFAULT 'Pending',
            created_at TIMESTAMP
        )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_withdrawals_user_id ON withdrawals(user_id)")

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS sub_admins (
            user_id INTEGER PRIMARY KEY,
            added_by INTEGER,
            created_at TIMESTAMP
        )
        """)

        default_settings = {
            "ig_2fa_price": "4.00",
            "ig_seed_price": "4.30",
            "fb_2fa_price": "5.00",
            "fb_cookies_price": "7.00",
            "gmail_task_price": "22.00",
            "ig_default_password": "RBKpass@05",
            "fb_default_password": "FBKpass@05",
            "gmail_default_password": "aass1122",
            "wd_bkash_active": "ON",
            "wd_nagad_active": "ON",
            "wd_usdt_active": "ON",
            "min_withdraw_bkash": "50.0",
            "max_withdraw_bkash": "5000.0",
            "min_withdraw_nagad": "50.0",
            "max_withdraw_nagad": "5000.0",
            "min_withdraw_usdt": "0.25",
            "max_withdraw_usdt": "100.0",
            "usdt_fee": "0.05",
            "admin_id": "8001997389",
            "support_handle": "https://t.me/TrustVaultMails_Owners",
            "official_channel_link": "https://t.me/TrustVaultMailsOfficial",
            "ref_bonus_percent": "10.0",
            "force_join_enabled": "OFF",
            "force_channels_list": "@TrustVaultMailsOfficial|https://t.me/TrustVaultMailsOfficial",
            "ig_2fa_active": "ON",
            "ig_seed_active": "ON",
            "fb_2fa_active": "ON",
            "fb_cookies_active": "ON",
            "gmail_task_active": "ON",
            "maintenance_mode": "OFF",
        }

        for key, val in default_settings.items():
            cursor.execute("INSERT OR IGNORE INTO admin_settings (key, value) VALUES (?, ?)", (key, str(val)))

        # One-time migration for the old mismatched Force Join default.
        cursor.execute("SELECT value FROM admin_settings WHERE key = ?", ("force_channels_list",))
        fj_row = cursor.fetchone()
        if fj_row and fj_row["value"] == "@TrustVaultMails_Owners|https://t.me/TrustVaultMailsOfficial":
            cursor.execute(
                "UPDATE admin_settings SET value = ? WHERE key = ?",
                ("@TrustVaultMailsOfficial|https://t.me/TrustVaultMailsOfficial", "force_channels_list")
            )

        # Security hardening: this copy belongs to the new owner only.
        # Force the configured owner/admin UID and remove every other sub-admin.
        cursor.execute(
            "UPDATE admin_settings SET value = ? WHERE key = ?",
            ("8001997389", "admin_id")
        )
        cursor.execute(
            "DELETE FROM sub_admins WHERE user_id != ?",
            (8001997389,)
        )
        conn.commit()
    logger.info("Database initialized successfully.")

# -----------------------------------------------------------------------------
# CENTRALIZED MULTILINGUAL TRANSLATION SYSTEM
# -----------------------------------------------------------------------------
MESSAGES = {
    "bn": {
        "btn_task": "💼 কাজ (Task)",
        "btn_wallet": "👛 ওয়ালেট (Wallet)",
        "btn_ref": "👥 আমার রেফারেল",
        "btn_leaderboard": "🏆 লিডারবোর্ড",
        "btn_withdraw": "💸 টাকা উত্তোলন (Withdraw)",
        "btn_support": "🎧 সাপোর্ট",
        "btn_lang": "🌐 ভাষা পরিবর্তন",
        "btn_admin_panel": "🟢 Admin Panel",
        "btn_main_menu": "🟢 মূল মেনু",
        "btn_cancel": "❌ বাতিল",
        "btn_instagram": "📸 ইনস্টাগ্রাম",
        "btn_facebook": "📘 ফেসবুক",
        "btn_gmail": "✉️ জিমেইল",
        "btn_ig_2fa": "📸 ইনস্টাগ্রাম ২এফএ",
        "btn_ig_seed": "📸 ইনস্টাগ্রাম সীড",
        "btn_fb_2fa": "📘 ফেসবুক ২এফএ",
        "btn_fb_cookies": "📘 ফেসবুক কুকিজ",
        "btn_gen_2fa": "🔑 2FA Code Generate করুন",
        "btn_acct_done": "✅ একাউন্ট খোলা শেষ",
        "btn_bkash": "🟢 বিকাশ (Bkash)",
        "btn_nagad": "🟠 নগদ (Nagad)",
        "btn_usdt": "🔵 USDT (BEP-20)",
        "btn_pending_tasks": "🟢 📩 পেন্ডিং টাস্ক",
        "btn_held_tasks": "🟢 ⌛ হোল্ড টাস্ক",
        "btn_task_search": "🟢 🔍 টাস্ক সার্চ",
        "btn_user_export": "🟢 📂 ইউজার এক্সপোর্ট",
        "btn_pwd_mgr": "🟢 🔑 পাসওয়ার্ড ম্যানেজ",
        "btn_price_mgr": "🟢 🏷️ প্রাইস ম্যানেজ",
        "btn_wd_reqs": "🟢 💸 উইথড্র রিকোয়েস্ট",
        "btn_sys_settings": "🟢 ⚙️ সিস্টেম সেটিংস",
        "btn_live_stats": "🟢 📊 লাইভ স্ট্যাটাস",
        "btn_broadcast": "🟢 📢 ব্রডকাস্ট",

        "btn_bulk_approve": "🟢 ✅ বাল্ক এপ্রুভ",
        "btn_bulk_reject": "🟢 ❌ বাল্ক রিজেক্ট",
        "btn_excel_export": "🟢 📊 এক্সেল শীট (Spreadsheet)",
        "btn_user_mgmt": "🟢 👥 ইউজার ম্যানেজমেন্ট",
        "btn_admin_mgmt": "🟢 👑 এডমিন কন্ট্রোল",

        "maintenance_msg": "⚙️ **বটটি বর্তমানে মেইনটেনেন্স মোডে রয়েছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।**",
        "banned_msg": "❌ আপনার অ্যাকাউন্টটি সিস্টেমে নিষিদ্ধ করা হয়েছে।",
        "force_join_msg": "📢 **বটটি ব্যবহার করতে আপনাকে আমাদের অফিশিয়াল চ্যানেলে যুক্ত থাকতে হবে:**\n\n📌 অনুগ্রহ করে নিচের সবকটি চ্যানেলে জয়েন করে **✅ ভেরিফাই করুন** বাটনে চাপ দিন:",
        "btn_join_channel": "🔗 জয়েন করুন Channel",
        "btn_verify": "✅ ভেরিফাই করুন",
        "force_join_success": "✅ **ধন্যবাদ! আপনার চ্যানেল ভেরিফিকেশন সফল হয়েছে।**",
        "force_join_fail": "❌ আপনি এখনো সবগুলো চ্যানেলে জয়েন করেননি! সব চ্যানেলে জয়েন করে পুনরায় চেষ্টা করুন।",
        "welcome_msg": "👋 **স্বাগতম, {name}!** 💎✨\n━━━━━━━━━━━━━━━━━━━━━━\nTrust Vault Mails Style প্ল্যাটফর্মে আপনাকে স্বাগতম।\nএখানে আপনি সহজ সোশ্যাল মিডিয়া অ্যাকাউন্ট ক্রিয়েশন এবং ভেরিফিকেশনের কাজ সম্পন্ন করে নিশ্চিত ইনকাম করতে পারবেন।\n━━━━━━━━━━━━━━━━━━━━━━\n👉 **কাজ শুরু করতে নিচের 💼 কাজ (Task) বাটনে চাপ দিন।**",
        "ref_new_user_joined": "🎉 **নতুন রেফারেল যোগ হয়েছে!**\n━━━━━━━━━━━━━━━━━━━━━━\n👤 **ইউজার:** {name}\n🆔 **ইউজার আইডি:** <code>{user_id}</code>\n\nকর্মী কাজ সম্পন্ন করলে আপনি পাবেন ১০% সরাসরি কমিশন!",
        "wd_all_disabled": "❌ বর্তমানে সকল পেমেন্ট মেথড বন্ধ রয়েছে। অনুগ্রহ করে পরে চেষ্টা করুন।",
        "task_cat_prompt": "📋 **অনুগ্রহ করে নিচের তালিকা থেকে আপনার পছন্দের কাজের ক্যাটাগরি নির্বাচন করুন:**",
        "task_off_ig": "❌ ইনস্টাগ্রামের কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_fb": "❌ ফেসবুকের কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_gmail": "❌ জিমেইলের কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_ig_2fa": "❌ ইনস্টাগ্রাম ২এফএ কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_ig_seed": "❌ ইনস্টাগ্রাম সীড কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_fb_2fa": "❌ ফেসবুক ২এফএ কাজ বর্তমানে বন্ধ রয়েছে।",
        "task_off_fb_cookies": "❌ ফেসবুক কুকিজ কাজ বর্তমানে বন্ধ রয়েছে।",
        "ig_menu_title": "📸 **ইনস্টাগ্রাম কাজ (Instagram Tasks):**\n━━━━━━━━━━━━━━━━━━━━━━\nনিচের যেকোনো একটি কাজের ধরণ নির্বাচন করুন:",
        "fb_menu_title": "📘 **ফেসবুক কাজ (Facebook Tasks):**\n━━━━━━━━━━━━━━━━━━━━━━\nনিচের যেকোনো একটি কাজের ধরণ নির্বাচন করুন:",
        "gmail_guidelines": "✉️ **জিমেইল টাস্ক গাইডলাইন (Gmail Task):**\n━━━━━━━━━━━━━━━━━━━━━━\n💵 **টাস্ক বোনাস:** ৳{price}\n🔑 **নির্ধারিত পাসওয়ার্ড (ট্যাপ করে কপি করুন):**\n<code>{fixed_pass}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n👉 ১. একটি নতুন জিমেইল অ্যাকাউন্ট তৈরি করে পাসওয়ার্ড হিসেবে ওপরের পাসওয়ার্ডটি দিন।\n👉 ২. কাজ শেষ হলে আপনার তৈরি করা **জিমেইল অ্যাড্রেসটি** নিচে লিখে পাঠান:",
        "ig_2fa_guidelines": "🔐 **ইনস্টাগ্রাম ২এফএ গাইডলাইন (Security Setup):**\n━━━━━━━━━━━━━━━━━━━━━━\n💵 **টাস্ক বোনাস:** ৳{price}\n👤 **ইউজারনেম:** <code>{username}</code>\n🔑 **পাসওয়ার্ড:** <code>{fixed_pass}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n👉 **ধাপ ১:** এই ইউজারনেম ও পাসওয়ার্ড দিয়ে একাউন্ট সেটআপ করুন।\n👉 **ধাপ ২:** টু-ফ্যাক্টর (2FA) সেটিংস চালু করে সেখান থেকে সিক্রেট কি কপি করুন।\n👉 **ধাপ ৩:** নিচে দেওয়া **🔑 2FA Code Generate** বাটনে চাপ দিয়ে ওটিপি নিয়ে অ্যাকাউন্ট অ্যাক্টিভেট করুন।",
        "ig_seed_guidelines": "📸 **ইনস্টাগ্রাম সীড ডেটা জমা দিন (৳{price}):**\n━━━━━━━━━━━━━━━━━━━━━━\nআপনার Instagram সীড ডেটা ফাইল (.xlsx) অথবা নিচের ফরম্যাটে টেক্সট পাঠান:\n<code>ইউজারনেম|পাসওয়ার্ড|সিক্রেটকি</code>\n\nফাইল বা টেক্সট পাঠানোর পর **'✅ জমা দিন (Done)'** বাটনে চাপ দিন।",
        "btn_done": "✅ জমা দিন (Done)",
        "file_received_msg": "📄 **ফাইল/ডেটা সফলভাবে গ্রহণ করা হয়েছে!**\n\nকাজটি রিভিউয়ের জন্য জমা দিতে নিচে **'✅ জমা দিন (Done)'** বাটনে চাপ দিন।",
        "no_file_or_data_err": "⚠️ **কোনো ফাইল বা ডেটা পাওয়া যায়নি!**\n\nপ্রথমে এক্সেল ফাইল (.xlsx) বা টেক্সট পাঠান, তারপর **'✅ জমা দিন (Done)'** বাটনে চাপ দিন।",
        "fb_2fa_guidelines": "📘 **ফেসবুক ২এফএ সেটআপ:**\n━━━━━━━━━━━━━━━━━━━━━━\n💵 **টাস্ক বোনাস:** ৳{price}\n🏷️ **প্রথম নাম:** <code>{first_name}</code>\n🏷️ **শেষ নাম:** <code>{last_name}</code>\n🔑 **পাসওয়ার্ড:** <code>{fixed_pass}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n👉 ১. অ্যাকাউন্টে ওপরের নামটি ব্যবহার করুন।\n👉 ২. টু-ফ্যাক্টর সেটিংস চালু করে সিক্রেট কি কপি করুন।\n👉 ৩. নিচে দেওয়া '🔑 2FA Code Generate করুন' বাটনে চাপ দিয়ে সিক্রেট কি পাঠান ও ওটিপি নিয়ে একাউন্ট জমা দিন।",
        "fb_cookies_guidelines_uid": "📘 **ফেসবুক কুকিজ টাস্ক:**\n━━━━━━━━━━━━━━━━━━━━━━\n💵 **টাস্ক বোনাস:** ৳{price}\n🏷️ **প্রথম নাম:** <code>{first_name}</code>\n🏷️ **শেষ নাম:** <code>{last_name}</code>\n🔑 **পাসওয়ার্ড:** <code>{fixed_pass}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n👉 **ধাপ ১:** ফেসবুক অ্যাকাউন্টের **UID** নিচে টাইপ করে পাঠান:",
        "fb_cookies_guidelines_str": "🍪 **এখন এই ফেসবুক অ্যাকাউন্টের Cookie String-টি কপি করে নিচে পেস্ট করুন:**",
        "otp_prompt_secret": "🔑 **আপনার 2FA Secret Key-টি প্রদান করুন:**\n\nঅ্যাপ থেকে প্রাপ্ত সিক্রেট কোডটি নিচে মেসেজ করুন:",
        "otp_invalid_secret": "❌ **ভুল সিক্রেট কোড!**\n\nদয়া করে সঠিক ৩২ অক্ষরের বেস৩২ সিক্রেট কি পাঠান:",
        "otp_verified_msg": "✅ **ওটিপি সিক্রেট কি ভেরিফাই হয়েছে!**\n━━━━━━━━━━━━━━━━━━━━━━\n🔐 **সিক্রেট কি:** <code>{secret}</code>\n🔢 **লাইভ OTP:** <code>{otp}</code> (ট্যাপ করে কপি করুন)\n━━━━━━━━━━━━━━━━━━━━━━\nএই কোডটি দিয়ে অ্যাকাউন্ট ভেরিফিকেশন সম্পন্ন করুন।\nসম্পন্ন হলে **'✅ একাউন্ট খোলা শেষ'** বাটনে চাপ দিন।",
        "otp_need_verify_first": "⚠️ **প্রথমে 2FA সিক্রেট কি ভেরিফাই করুন!**\n\n'🔑 2FA Code Generate করুন' বাটনে চাপ দিন।",
        "enter_username_prompt": "👤 **আপনার অ্যাকাউন্টের ইউজারনেম/ইমেইল/UID টাইপ করে পাঠান:**",
        "task_submitted_success": "🎉 <b>টাস্ক সফলভাবে জমা দেওয়া হয়েছে!</b> ✨\n━━━━━━━━━━━━━━━━━━━━━━\n🆔 <b>টাস্ক আইডি:</b> #{sub_id}\n📱 <b>টাইপ:</b> {task_display}\n👤 <b>ইউজারনেম:</b> <code>{username}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n🔍 <b>আপনার তথ্য সফলভাবে পাওয়া গেছে!</b>\n\nআগামী ২৪ থেকে ৪৮ ঘণ্টার মধ্যে আপনার একাউন্ট চেক করে ব্যালেন্স যোগ করে দেওয়া হবে।\n\nএডমিন রিভিউ করে দ্রুত আপনার ওয়ালেটে ব্যালেন্স যুক্ত করে দেবে।\n\nধন্যবাদ!",
        "ig_seed_submitted_success": "🎉 **ইনস্টাগ্রাম সীড ডেটা জমা দেওয়া হয়েছে!**\n\n🆔 **টাস্ক আইডি:** #{sub_id}\n🔍 **আপনার তথ্য সফলভাবে পাওয়া গেছে!**\n\nআগামী ২৪ থেকে ৪৮ ঘণ্টার মধ্যে আপনার একাউন্ট চেক করে ব্যালেন্স যোগ করে দেওয়া হবে।\n\nএডমিন রিভিউ করে দ্রুত আপনার ওয়ালেটে ব্যালেন্স যুক্ত করে দেবে।\n\nধন্যবাদ!",
        "fb_cookies_submitted_success": "🎉 **ফেসবুক কুকিজ টাস্ক জমা দেওয়া হয়েছে!**\n\n🆔 **টাস্ক আইডি:** #{sub_id}\n🔍 **আপনার তথ্য সফলভাবে পাওয়া গেছে!**\n\nআগামী ২৪ থেকে ৪৮ ঘণ্টার মধ্যে আপনার একাউন্ট চেক করে ব্যালেন্স যোগ করে দেওয়া হবে।\n\nএডমিন রিভিউ করে দ্রুত আপনার ওয়ালেটে ব্যালেন্স যুক্ত করে দেবে।\n\nধন্যবাদ!",
        "gmail_submitted_success": "🎉 **জিমেইল টাস্ক জমা দেওয়া হয়েছে!**\n\n🆔 **টাস্ক আইডি:** #{sub_id}\n🔍 **আপনার তথ্য সফলভাবে পাওয়া গেছে!**\n\nআগামী ২৪ থেকে ৪৮ ঘণ্টার মধ্যে আপনার একাউন্ট চেক করে ব্যালেন্স যোগ করে দেওয়া হবে।\n\nএডমিন রিভিউ করে দ্রুত আপনার ওয়ালেটে ব্যালেন্স যুক্ত করে দেবে।\n\nধন্যবাদ!",
        "operation_cancelled": "❌ কাজটি বাতিল করা হয়েছে।",
        "returning_main_menu": "🔙 মূল মেনুতে ফিরে যাচ্ছেন...",

        "wallet_text": "👛 **আপনার অ্যাকাউন্ট ওয়ালেট** 💎\n━━━━━━━━━━━━━━━━━━━━━━\n💰 **বর্তমান ব্যালেন্স:** ৳{balance:.2f}\n━━━━━━━━━━━━━━━━━━━━━━\n💸 **পেন্ডিং উইথড্র:** ৳{pending_wd:.2f}\n💰 **মোট আয় (Lifetime):** ৳{lifetime:.2f}\n👥 **রেফারেল আয়:** ৳{ref_earnings:.2f}\n━━━━━━━━━━━━━━━━━━━━━━\n✅ **সম্পন্ন কাজ:** {completed} টি\n⏳ **রিভিউতে আছে:** {pending} টি\n⌛ **হোল্ডে আছে:** {held} টি\n❌ **রিজেক্টেড:** {rejected} টি",
        "wd_gateway_prompt": "💸 **উইথড্রয়াল পেমেন্ট গেটওয়ে নির্বাচন করুন:**\n━━━━━━━━━━━━━━━━━━━━━━\nপছন্দের মেথড নির্বাচন করুন:",
        "wd_invalid_method": "❌ সঠিক উইথড্রয়াল মেথড নির্বাচন করুন।",
        "wd_enter_account_usdt": "📱 **আপনার USDT (BEP-20) ওয়ালেট অ্যাড্রেস লিখুন:**",
        "wd_enter_account_mfs": "📱 **আপনার {method} অ্যাকাউন্ট নম্বর লিখুন (১১-১৪ ডিজিট):**",
        "wd_invalid_account": "❌ ভুল অ্যাকাউন্ট নম্বর! সঠিক ১১-১৪ ডিজিটের নম্বর দিন।",
        "wd_enter_amount": "💵 **উইথড্র করার পরিমাণ লিখুন:**\n━━━━━━━━━━━━━━━━━━━━━━\n💰 আপনার বর্তমান ব্যালেন্স: ৳{balance:.2f}\n🔻 সর্বনিম্ন সীমা: ৳{min_limit}\n🔺 সর্বোচ্চ সীমা: ৳{max_limit}",
        "wd_invalid_amount_number": "❌ সঠিক সংখ্যা লিখুন (যেমন: ১০০, ২০০.৫):",
        "wd_min_limit_err": "❌ সর্বনিম্ন উইথড্রসীমা ৳{min_limit}! পুনরায় চেষ্টা করুন।",
        "wd_max_limit_err": "❌ সর্বোচ্চ উইথড্রসীমা ৳{max_limit}! পুনরায় চেষ্টা করুন।",
        "wd_insufficient_bal": "❌ আপনার অ্যাকাউন্টে পর্যাপ্ত ব্যালেন্স নেই! পুনরায় চেষ্টা করুন।",
        "wd_usdt_fee_err": "❌ ফি বাদ দেওয়ার পরে পরিমাণ ০ এর কম! বড় পরিমাণ লিখুন।",
        "wd_success_msg": "✅ **পেমেন্ট রিকোয়েস্ট সফলভাবে সাবমিট হয়েছে!** ✨\n━━━━━━━━━━━━━━━━━━━━━━\n🆔 **উইথড্র আইডি:** #{wd_id}\n💵 **পরিমাণ:** ৳{amount:.2f}{fee_text}\n📱 **মেথড:** {method} ({number})\n━━━━━━━━━━━━━━━━━━━━━━\n💰 আপনার অ্যাকাউন্ট থেকে ৳{amount:.2f} কেটে নেওয়া হয়েছে।\n⏳ স্ট্যাটাস: পেন্ডিং (১২-২৪ ঘণ্টার মধ্যে রিভিউ সম্পন্ন হবে)",

        "ref_text": "👥 **আপনার ব্যক্তিগত রেফারেল প্রোগ্রাম** 💎\n━━━━━━━━━━━━━━━━━━━━━━\nআপনার রেফারেল লিঙ্ক ব্যবহার করে বন্ধুদের আমন্ত্রণ জানান। আপনার আমন্ত্রিত কর্মী যতগুলো কাজ সম্পন্ন ও এপ্রুভ করাবে, প্রতিটি কাজের ওপর আপনি পাবেন সরাসরি **{ref_percent}% কমিশন** আপনার ওয়ালেটে।\n\n🔗 **আপনার রেফারেল লিঙ্ক:**\n<code>{ref_link}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n📊 **রেফারেল ট্র্যাকার:**\n• মোট আমন্ত্রিত কর্মী: <code>{ref_count}</code> জন\n• রেফারেল বোনাস থেকে আয়: <code>৳{ref_earnings:.2f}</code>",
        "leaderboard_title": "🏆 **শীর্ষ ১০ উপার্জনকারী কর্মী (Leaderboard):** 👑\n━━━━━━━━━━━━━━━━━━━━━━\n",
        "no_leaderboard_data": "কোনো ডেটা এখনো পাওয়া যায়নি।",
        "support_text": "🎧 **গ্রাহক সেবা কেন্দ্র (Help Zone)**\n━━━━━━━━━━━━━━━━━━━━━━\n👤 সম্মানিত মেম্বার, আপনার যেকোনো প্রশ্নের জন্য আমাদের সাপোর্ট টিমের সাথে যোগাযোগ করুন।\n📌 অযথা মেসেজ দেওয়া থেকে বিরত থাকুন।",
        "btn_admin_support": "💎 এডমিন সাপোর্ট",
        "btn_official_channel": "🔗 অফিসিয়াল চ্যানেল",
        "lang_select_prompt": "🌐 **ভাষা নির্বাচন করুন / Select Language:**\n\nনিচের বাটন থেকে আপনার পছন্দের ভাষা নির্বাচন করুন:",
        "lang_set_bn": "✅ বটের ভাষা সফলভাবে **বাংলা** করা হয়েছে।",
        "lang_set_en": "✅ Bot language set to **English**.",

        "user_task_approved": "✅ <b>অভিনন্দন!</b> আপনার [{task_display}] কাজ (<code>{username}</code>) এপ্রুভ হয়েছে। ৳{reward:.2f} ব্যালেন্সে যোগ হয়েছে।",
        "user_task_rejected": "❌ <b>আপনার [{task_display}] কাজ ({username}) রিজেক্ট করা হয়েছে।</b>\n📋 <b>কারণ:</b> {reason}",
        "user_task_held": "⏳ <b>আপনার টাস্ক #{sub_id} হোল্ডে রাখা হয়েছে।</b>\n\n<b>কারণ:</b>\nআপনার কাজটি ম্যানুয়াল রিভিউর অধীনে রয়েছে।",
        "user_ref_commission": "🎉 <b>রেফারেল কমিশন যুক্ত হয়েছে!</b>\n\nআপনার রেফারকৃত কর্মী একটি কাজ সম্পন্ন করায় আপনি <b>৳{bonus:.2f}</b> কমিশন পেয়েছেন।",
        "user_wd_approved": "✅ <b>আপনার উইথড্র রিকোয়েস্ট #{wd_id} সফলভাবে সম্পন্ন হয়েছে!</b>\n💵 <b>৳{amount:.2f}</b> আপনার <code>{number}</code> নম্বরে পাঠানো হয়েছে।",
        "user_wd_rejected": "❌ <b>আপনার উইথড্র রিকোয়েস্ট #{wd_id} রিজেক্ট করা হয়েছে।</b>\n💵 <b>৳{amount:.2f}</b> আপনার ওয়ালেটে রিফান্ড (ফেরত) করা হয়েছে।",

        "admin_dash_title": "👑 **Trust Vault Mails Style: অ্যাডমিন প্যানেল** 👑\n━━━━━━━━━━━━━━━━━━━━━━\n👥 **মোট ইউজার:** {total_users} জন\n⏳ **পেন্ডিং কাজ:** {pending_tasks} টি\n⌛ **হোল্ড কাজ:** {held_tasks} টি\n💸 **পেন্ডিং উইথড্র:** {pending_wd} টি\n━━━━━━━━━━━━━━━━━━━━━━\nযেকোনো বিষয় কনফিগার বা ভেরিফাই করতে নিচের মেনুগুলো ব্যবহার করুন।",
        "admin_no_pending_tasks": "✅ কোনো পেন্ডিং টাস্ক নেই।",
        "admin_pending_tasks_title": "📩 **পেন্ডিং টাস্ক লিস্ট (সর্বশেষ ১০টি):**\n\n",
        "admin_no_held_tasks": "⏳ কোনো হোল্ড টাস্ক নেই।",
        "admin_held_tasks_title": "⌛ <b>হোল্ড টাস্ক লিস্ট:</b>\n\n",
        "admin_search_prompt_msg": "🔍 **টাস্ক সার্চ:**\nইউজারনেম, ফেসবুক UID, জিমেইল, অথবা টাস্ক আইডি লিখুন:",
        "admin_export_prompt_msg": "📂 **ইউজার এক্সপোর্ট:**\nযে ইউজারের ডেটা চান তার টেলিগ্রাম ইউজার আইডি লিখুন:",
        "admin_pwd_mgr_title": "🔑 **পাসওয়ার্ড ম্যানেজ**\n━━━━━━━━━━━━━━━━━━━━━━\n📸 IG: <code>{ig_pass}</code>\n📘 FB: <code>{fb_pass}</code>\n✉️ Gmail: <code>{gmail_pass}</code>\n━━━━━━━━━━━━━━━━━━━━━━\nনতুন পাসওয়ার্ড সেট করতে নিচের বাটন ব্যবহার করুন।",
        "admin_price_mgr_title": "🏷️ **টাস্ক প্রাইস ম্যানেজার (Task Price Management)**\n━━━━━━━━━━━━━━━━━━━━━━\n📸 IG 2FA: <code>৳{ig_2fa}</code>\n📸 IG Seed: <code>৳{ig_seed}</code>\n📘 FB 2FA: <code>৳{fb_2fa}</code>\n📘 FB Cookies: <code>৳{fb_cookies}</code>\n✉️ Gmail: <code>৳{gmail}</code>\n👥 Ref Bonus: <code>{ref_bonus}%</code>\n━━━━━━━━━━━━━━━━━━━━━━\nনতুন দাম নির্ধারণ করতে নিচের বাটনে চাপ দিন:",
        "admin_no_pending_wds": "✅ কোনো পেন্ডিং উইথড্র রিকোয়েস্ট নেই।",
        "admin_pending_wds_title": "💸 **পেন্ডিং উইথড্র রিকোয়েস্ট:**\n\n",
        "admin_sys_settings_title": "⚙️ **সিস্টেম সেটিংস (ON/OFF Toggles & Controls)**\n━━━━━━━━━━━━━━━━━━━━━━\n🔒 Force Join Status: {force_join}\n🛠️ Maintenance Mode: {maint}\n📸 IG 2FA: {ig_2fa} | IG Seed: {ig_seed}\n📘 FB 2FA: {fb_2fa} | FB Cookies: {fb_cookies}\n✉️ Gmail: {gmail}\n🟢 Bkash WD: {bkash_wd} (Min: {min_bkash} | Max: {max_bkash})\n🟠 Nagad WD: {nagad_wd} (Min: {min_nagad} | Max: {max_nagad})\n🔵 USDT WD: {usdt_wd} (Min: {min_usdt} | Max: {max_usdt})\n━━━━━━━━━━━━━━━━━━━━━━\nটগল বা পরিবর্তন করতে নিচের বাটন ব্যবহার করুন।",
        "admin_live_stats_title": "📊 **লাইভ স্ট্যাটাস** ⚡\n━━━━━━━━━━━━━━━━━━━━━━\n👥 মোট ইউজার: {total_users}\n⏳ পেন্ডিং টাস্ক: {pending_tasks}\n⌛ হোল্ড টাস্ক: {held_tasks}\n✅ এপ্রুভড টাস্ক: {approved_tasks}\n❌ রিজেক্টেড টাস্ক: {rejected_tasks}\n💸 পেন্ডিং উইথড্র: {pending_wd}\n💰 মোট ইউজার ব্যালেন্স: ৳{total_balance:.2f}",
        "admin_broadcast_prompt_msg": "𝚁𝙱𝙺 𝙰𝙿𝙿 𝚂𝚃𝙾𝚁𝙴: 📢 Broadcast Message পাঠান (All Media Supported)\n━━━━━━━━━━━━━━━━━━━━━━\nএখানে পছন্দমতো যেকোনো **Text, Photo 🖼️, Video 🎥, Document 📄, Voice 🎤, Audio 🎵** সেন্ড করুন। ক্যাপশন থাকলে ক্যাপশনসহ ইউজারদের কাছে অরিজিনাল মেসেজ হিসেবে যাবে।",
        "admin_search_no_results": "❌ কোনো টাস্ক পাওয়া যায়নি।",
        "admin_search_results_title": "🔍 **সার্চ রেজাল্ট:**\n\n",
        "admin_export_invalid_id": "❌ সঠিক ইউজার আইডি দিন।",
        "admin_export_no_tasks": "❌ এই ইউজারের কোনো টাস্ক নেই।",
        "admin_broadcast_done": "✅ **Broadcast Completed** ✨\n━━━━━━━━━━━━━━━━━━━━━━\n👥 **Total Users:** {total}\n📨 **Message Type:** {msg_type}\n✅ **Successfully Sent:** {count}\n❌ **Failed:** {failed}\n📈 **Success Rate:** {success_rate}%\n⏱️ **Total Time:** {time_taken}s\n━━━━━━━━━━━━━━━━━━━━━━\n📊 **Failure Breakdown:**\n• 🚫 Bot Blocked: {blocked}\n• 🗑️ Account Deleted / Chat Not Found: {deleted}\n• ⚠️ Other Errors: {other}",
        "admin_rej_custom_prompt": "📝 **রিজেক্ট করার কাস্টম কারণটি লিখে পাঠান:**",
        "btn_admin_menu": "🟢 Admin Menu",
        "admin_bulk_app_prompt": "✅ **বাল্ক এপ্রুভ (Bulk Approve):**\n━━━━━━━━━━━━━━━━━━━━━━\nএখানে একসাথে ৫০-২০০টি Username / UID / Gmail পেস্ট করে সেন্ড করুন (প্রতি লাইনে একটি করে অথবা কমা/স্পেস দিয়ে):\n\nবট অটোম্যাটিক্যালি এগুলো চেক করে **শুধুমাত্র ইনপুট দেওয়া টাস্কগুলো এপ্রুভ** করবে ও ইউজারকে ওয়ালেট ব্যালেন্সসহ নোটিফিকেশন পাঠাবে।",
        "admin_bulk_rej_prompt": "❌ **বাল্ক রিজেক্ট (Bulk Reject):**\n━━━━━━━━━━━━━━━━━━━━━━\nএখানে একসাথে ৫০-২০০টি Username / UID / Gmail পেস্ট করে সেন্ড করুন (প্রতি লাইনে একটি করে অথবা কমা/স্পেস দিয়ে):\n\nবট অটোম্যাটিক্যালি এগুলো চেক করে **শুধুমাত্র ইনপুট দেওয়া টাস্কগুলো রিজেক্ট** করবে ও সংশ্লিষ্ট ইউজারকে নোটিফিকেশন পাঠাবে।",
    },
    "en": {
        "btn_task": "💼 Task",
        "btn_wallet": "👛 Wallet",
        "btn_ref": "👥 My Referrals",
        "btn_leaderboard": "🏆 Leaderboard",
        "btn_withdraw": "💸 Withdraw",
        "btn_support": "🎧 Support",
        "btn_lang": "🌐 Language",
        "btn_admin_panel": "🟢 Admin Panel",
        "btn_main_menu": "🟢 Main Menu",
        "btn_cancel": "❌ Cancel",
        "btn_instagram": "📸 Instagram",
        "btn_facebook": "📘 Facebook",
        "btn_gmail": "✉️ Gmail",
        "btn_ig_2fa": "📸 Instagram 2FA",
        "btn_ig_seed": "📸 Instagram Seed",
        "btn_fb_2fa": "📘 Facebook 2FA",
        "btn_fb_cookies": "📘 Facebook Cookies",
        "btn_gen_2fa": "🔑 Generate 2FA Code",
        "btn_acct_done": "✅ Account Creation Done",
        "btn_bkash": "🟢 Bkash",
        "btn_nagad": "🟠 Nagad",
        "btn_usdt": "🔵 USDT (BEP-20)",
        "btn_pending_tasks": "🟢 📩 Pending Tasks",
        "btn_held_tasks": "🟢 ⌛ Held Tasks",
        "btn_task_search": "🟢 🔍 Task Search",
        "btn_user_export": "🟢 📂 User Export",
        "btn_pwd_mgr": "🟢 🔑 Password Manager",
        "btn_price_mgr": "🟢 🏷️ Price Manager",
        "btn_wd_reqs": "🟢 💸 Withdraw Requests",
        "btn_sys_settings": "🟢 ⚙️ System Settings",
        "btn_live_stats": "🟢 📊 Live Stats",
        "btn_broadcast": "🟢 📢 Broadcast",

        "btn_bulk_approve": "🟢 ✅ Approve All",
        "btn_bulk_reject": "🟢 ❌ Reject All",
        "btn_excel_export": "🟢 📊 Spreadsheet UI",
        "btn_user_mgmt": "🟢 👥 User Management",
        "btn_admin_mgmt": "🟢 👑 Admin Control",

        "maintenance_msg": "⚙️ **The bot is currently under maintenance. Please try again later.**",
        "banned_msg": "❌ Your account has been banned from the system.",
        "force_join_msg": "📢 **To use this bot, you must join our official channels:**\n\nPlease join all channels below and press the **✅ Verify** button:",
        "btn_join_channel": "🔗 Join Channel",
        "btn_verify": "✅ Verify",
        "force_join_success": "✅ **Thank you! Your channel verification was successful.**",
        "force_join_fail": "❌ You haven't joined all channels yet! Please join all channels and try again.",
        "welcome_msg": "👋 **Welcome, {name}!** 💎✨\n━━━━━━━━━━━━━━━━━━━━━━\nWelcome to Trust Vault Mails Style Tasking Bot.\nHere you can earn money by completing simple social account creation and verification tasks.\n━━━━━━━━━━━━━━━━━━━━━━\n👉 **Press the 💼 Task button below to start working.**",
        "ref_new_user_joined": "🎉 **New Referral Joined!**\n━━━━━━━━━━━━━━━━━━━━━━\n👤 **User:** {name}\n🆔 **User ID:** <code>{user_id}</code>\n\nYou will earn 10% lifetime commission on their completed tasks!",
        "wd_all_disabled": "❌ All withdrawal gateways are currently disabled. Please try again later.",
        "task_cat_prompt": "📋 **Please select a task category from the list below:**",
        "task_off_ig": "❌ Instagram tasks are currently disabled.",
        "task_off_fb": "❌ Facebook tasks are currently disabled.",
        "task_off_gmail": "❌ Gmail tasks are currently disabled.",
        "task_off_ig_2fa": "❌ Instagram 2FA tasks are currently disabled.",
        "task_off_ig_seed": "❌ Instagram Seed tasks are currently disabled.",
        "task_off_fb_2fa": "❌ Facebook 2FA tasks are currently disabled.",
        "task_off_fb_cookies": "❌ Facebook Cookies tasks are currently disabled.",
        "ig_menu_title": "📸 **Instagram Tasks:**\n\nSelect one of the options below:",
        "fb_menu_title": "📘 **Facebook Tasks:**\n\nSelect one of the options below:",
        "gmail_guidelines": "✉️ **Gmail Task Guidelines:**\n\n💵 **Task Bonus:** ৳{price}\n🔑 **Password (Tap to copy):**\n<code>{fixed_pass}</code>\n\n👉 Create a new Gmail account using the password above.\nOnce completed, reply with your created **Gmail address** below:",
        "ig_2fa_guidelines": "🔐 **Security Setup Guidelines:**\n💵 **Task Bonus:** ৳{price}\n👤 **Username:** <code>{username}</code>\n🔑 **Assigned Password:** <code>{fixed_pass}</code>\n\n👉 **Step 1:** Set up the account using this password.\n👉 **Step 2:** Enable 2FA settings and copy the Secret Key.\n👉 **Step 3:** Press the Generate OTP button below to activate the account with OTP.",
        "ig_seed_guidelines": "📸 **Submit Instagram Seed Data (৳{price}):**\n\nSend your Instagram Seed data file (.xlsx) or text in format:\n<code>username|password|secretkey</code>\n\nAfter sending file or text, press **'✅ Submit (Done)'**.",
        "btn_done": "✅ Submit (Done)",
        "file_received_msg": "📄 **File/Data received successfully!**\n\nPress **'✅ Submit (Done)'** below to send it to the admin.",
        "no_file_or_data_err": "⚠️ **No file or data found!**\n\nPlease send an Excel file (.xlsx) or text first, then press **'✅ Submit (Done)'**.",
        "fb_2fa_guidelines": "📘 **Facebook 2FA Setup:**\n\n💵 **Task Bonus:** ৳{price}\n🏷️ **First Name:** <code>{first_name}</code>\n🏷️ **Last Name:** <code>{last_name}</code>\n🔑 **Password:** <code>{fixed_pass}</code>\n\n👉 1. Use the name above on the account.\n👉 2. Turn on 2FA settings and copy the Secret Key.\n👉 3. Press '🔑 Generate 2FA Code' below, send the secret key, and activate the account using OTP.",
        "fb_cookies_guidelines_uid": "📘 **Facebook Cookies Task:**\n\n💵 **Task Bonus:** ৳{price}\n🏷️ **First Name:** <code>{first_name}</code>\n🏷️ **Last Name:** <code>{last_name}</code>\n🔑 **Password:** <code>{fixed_pass}</code>\n\n**Step 1:** Send the Facebook account **UID**:",
        "fb_cookies_guidelines_str": "🍪 **Now copy and paste the Cookie string of this Facebook account:**",
        "otp_prompt_secret": "🔑 **Provide your 2FA Secret Key:**\n\nSend the secret code obtained from the app here:",
        "otp_invalid_secret": "❌ **Invalid Secret Code!**\n\nPlease send a valid Base32 Secret Key:",
        "otp_verified_msg": "✅ **OTP Secret Key verified!**\n\n🔐 **Secret Key:** <code>{secret}</code>\n🔢 **Live OTP:** <code>{otp}</code> (Tap to copy)\n\nComplete account verification using this code.\nWhen finished, press the '✅ Account Creation Done' button.",
        "otp_need_verify_first": "⚠️ **Please verify your 2FA Secret Key first!**\n\nPress the '🔑 Generate 2FA Code' button.",
        "enter_username_prompt": "👤 **Type and send your account Username/Email/UID:**",
        "task_submitted_success": "🎉 <b>Task submitted successfully!</b> ✨\n━━━━━━━━━━━━━━━━━━━━━━\n🆔 <b>Task ID:</b> #{sub_id}\n📱 <b>Type:</b> {task_display}\n👤 <b>Username:</b> <code>{username}</code>\n━━━━━━━━━━━━━━━━━━━━━━\n🔍 <b>Your information has been received successfully!</b>\n\nYour account will be checked and the balance will be added within 24–48 hours.\n\nThe admin will review your submission and credit your wallet as soon as possible.\n\nThank you!",
        "ig_seed_submitted_success": "🎉 **Instagram Seed data submitted!**\n\n🆔 **Task ID:** #{sub_id}\n🔍 **Your information has been received successfully!**\n\nYour account will be checked and the balance will be added within 24–48 hours.\n\nThe admin will review your submission and credit your wallet as soon as possible.\n\nThank you!",
        "fb_cookies_submitted_success": "🎉 **Facebook Cookies task submitted!**\n\n🆔 **Task ID:** #{sub_id}\n🔍 **Your information has been received successfully!**\n\nYour account will be checked and the balance will be added within 24–48 hours.\n\nThe admin will review your submission and credit your wallet as soon as possible.\n\nThank you!",
        "gmail_submitted_success": "🎉 **Gmail task submitted!**\n\n🆔 **Task ID:** #{sub_id}\n🔍 **Your information has been received successfully!**\n\nYour account will be checked and the balance will be added within 24–48 hours.\n\nThe admin will review your submission and credit your wallet as soon as possible.\n\nThank you!",
        "operation_cancelled": "❌ Operation cancelled.",
        "returning_main_menu": "🔙 Returning to Main Menu...",

        "wallet_text": "💵 **Your Balance**\n━━━━━━━━━━━━━━━━━━━━━━\n💰 **Current Balance:** ৳{balance:.2f}\n━━━━━━━━━━━━━━━━━━━━━━\n💸 **Pending Withdrawal:** ৳{pending_wd:.2f}\n💰 **Lifetime Earnings:** ৳{lifetime:.2f}\n👥 **Referral Earnings:** ৳{ref_earnings:.2f}\n━━━━━━━━━━━━━━━━━━━━━━\n✅ **Completed Tasks:** {completed}\n⏳ **Under Review:** {pending}\n⌛ **On Hold:** {held}\n❌ **Rejected:** {rejected}",
        "wd_gateway_prompt": "💸 **Select Withdrawal Payment Gateway:**\n\nChoose payment method:",
        "wd_invalid_method": "❌ Please select a valid withdrawal method.",
        "wd_enter_account_usdt": "📱 **Enter your USDT (BEP-20) wallet address:**",
        "wd_enter_account_mfs": "📱 **Enter your {method} account number (11-14 digits):**",
        "wd_invalid_account": "❌ Invalid account number! Enter a valid 11-14 digit number.",
        "wd_enter_amount": "💵 **Enter withdrawal amount:**\n\n💰 Current Balance: ৳{balance:.2f}\n🔻 Minimum Limit: ৳{min_limit}\n🔺 Maximum Limit: ৳{max_limit}",
        "wd_invalid_amount_number": "❌ Enter a valid number (e.g. 100, 200.5):",
        "wd_min_limit_err": "❌ Minimum withdrawal limit is ৳{min_limit}! Try again.",
        "wd_max_limit_err": "❌ Maximum withdrawal limit is ৳{max_limit}! Try again.",
        "wd_insufficient_bal": "❌ Insufficient balance in your account! Try again.",
        "wd_usdt_fee_err": "❌ Amount after fee deduction is less than 0! Enter a larger amount.",
        "wd_success_msg": "✅ **Payment request submitted successfully!**\n\n🆔 **Withdrawal ID:** #{wd_id}\n💵 **Amount:** ৳{amount:.2f}{fee_text}\n📱 **Method:** {method} ({number})\n💰 ৳{amount:.2f} has been deducted from your current balance.\n⏳ Status: Pending (Review will be completed within 12-24 hours)",

        "ref_text": "👥 **My Referral Program**\n\nInvite your friends using your referral link. You will earn a direct **{ref_percent}% commission** on EVERY task your referred worker completes and gets approved.\n\n🔗 **Your Personal Referral Link:**\n<code>{ref_link}</code>\n\n📊 **Your Referral Tracker:**\n• Total Invited Workers: <code>{ref_count}</code>\n• Total Referral Earnings: <code>৳{ref_earnings:.2f}</code>",
        "leaderboard_title": "🏆 **Top 10 Earners:**\n\n",
        "no_leaderboard_data": "No data available yet.",
        "support_text": "↗️ **Customer Support Center**\n\n👤 Dear Member, for any issue please contact our support team.\n📌 Avoid sending unnecessary messages.",
        "btn_admin_support": "💎 Admin Support",
        "btn_official_channel": "🔗 Official Channel",
        "lang_select_prompt": "🌐 **Select Language:**\n\nChoose your preferred language from the buttons below:",
        "lang_set_bn": "✅ বটের ভাষা সফলভাবে **বাংলা** করা হয়েছে।",
        "lang_set_en": "✅ Bot language set to **English**.",

        "user_task_approved": "✅ <b>Congratulations!</b> Your [{task_display}] task (<code>{username}</code>) has been approved. ৳{reward:.2f} added to your balance.",
        "user_task_rejected": "❌ <b>Your [{task_display}] task ({username}) was rejected.</b>\n📋 <b>Reason:</b> {reason}",
        "user_task_held": "⏳ <b>Your task #{sub_id} has been placed on Hold.</b>\n\n<b>Reason:</b>\nYour task is under manual review.",
        "user_ref_commission": "🎉 <b>Referral Commission Added!</b>\n\nYou received <b>৳{bonus:.2f}</b> commission for a task completed by your referred worker.",
        "user_wd_approved": "✅ <b>Your withdrawal request #{wd_id} has been approved successfully!</b>\n💵 <b>৳{amount:.2f}</b> sent to your <code>{number}</code> account.",
        "user_wd_rejected": "❌ <b>Your withdrawal request #{wd_id} was rejected.</b>\n💵 <b>৳{amount:.2f}</b> refunded back to your wallet.",

        "admin_dash_title": "👑 **Trust Vault Mails Style: Admin Panel**\n\n👥 **Total Users:** {total_users}\n⏳ **Pending Tasks:** {pending_tasks}\n⌛ **Hold Tasks:** {held_tasks}\n💸 **Pending Withdrawals:** {pending_wd}\n\nUse the menus below to configure or verify anything.",
        "admin_no_pending_tasks": "✅ No pending tasks.",
        "admin_pending_tasks_title": "📩 **Pending Tasks List (Latest 10):**\n\n",
        "admin_no_held_tasks": "⏳ No held tasks.",
        "admin_held_tasks_title": "⌛ <b>Held Tasks List:</b>\n\n",
        "admin_search_prompt_msg": "🔍 **Task Search:**\nType Username, Facebook UID, Gmail, or Task ID:",
        "admin_export_prompt_msg": "📂 **User Export:**\nType Telegram User ID of the user you want data for:",
        "admin_pwd_mgr_title": "🔑 **Password Manager**\n\n📸 IG: <code>{ig_pass}</code>\n📘 FB: <code>{fb_pass}</code>\n✉️ Gmail: <code>{gmail_pass}</code>\n\nUse buttons below to set new passwords.",
        "admin_price_mgr_title": "🏷️ **Task Price Manager**\n\n📸 IG 2FA: <code>৳{ig_2fa}</code>\n📸 IG Seed: <code>৳{ig_seed}</code>\n📘 FB 2FA: <code>৳{fb_2fa}</code>\n📘 FB Cookies: <code>৳{fb_cookies}</code>\n✉️ Gmail: <code>৳{gmail}</code>\n👥 Ref Bonus: <code>{ref_bonus}%</code>\n\nUse buttons below to set new prices:",
        "admin_no_pending_wds": "✅ No pending withdrawal requests.",
        "admin_pending_wds_title": "💸 **Pending Withdrawal Requests:**\n\n",
        "admin_sys_settings_title": "⚙️ **System Settings (ON/OFF Toggles)**\n\n🔒 Force Join: {force_join}\n🛠️ Maintenance: {maint}\n📸 IG 2FA: {ig_2fa} | IG Seed: {ig_seed}\n📘 FB 2FA: {fb_2fa} | FB Cookies: {fb_cookies}\n✉️ Gmail: {gmail}\n🟢 Bkash WD: {bkash_wd} (Min: {min_bkash} | Max: {max_bkash})\n🟠 Nagad WD: {nagad_wd} (Min: {min_nagad} | Max: {max_nagad})\n🔵 USDT WD: {usdt_wd} (Min: {min_usdt} | Max: {max_usdt})\n\nUse buttons below to toggle or edit settings.",
        "admin_live_stats_title": "📊 **Live Stats**\n\n👥 Total Users: {total_users}\n⏳ Pending Tasks: {pending_tasks}\n⌛ Hold Tasks: {held_tasks}\n✅ Approved Tasks: {approved_tasks}\n❌ Rejected Tasks: {rejected_tasks}\n💸 Pending Withdrawals: {pending_wd}\n💰 Total User Balance: ৳{total_balance:.2f}",
        "admin_broadcast_prompt_msg": "𝚁𝙱𝙺 𝙰𝙿𝙿 𝚂𝚃𝙾𝚁𝙴: 📢 Broadcast Message পাঠান (All Media Supported)\n━━━━━━━━━━━━━━━━━━━━━━\nSend any **Text, Photo 🖼️, Video 🎥, Document 📄, Voice 🎤, Audio 🎵**. Caption will be preserved.",
        "admin_search_no_results": "❌ No tasks found.",
        "admin_search_results_title": "🔍 **Search Results:**\n\n",
        "admin_export_invalid_id": "❌ Enter a valid User ID.",
        "admin_export_no_tasks": "❌ This user has no tasks.",
        "admin_broadcast_done": "✅ **Broadcast Completed**\n\n👥 **Total Users:** {total}\n📨 **Message Type:** {msg_type}\n✅ **Successfully Sent:** {count}\n❌ **Failed:** {failed}\n📈 **Success Rate:** {success_rate}%\n⏱️ **Total Time:** {time_taken}s\n\n📊 **Failure Breakdown:**\n• 🚫 Bot Blocked: {blocked}\n• 🗑️ Account Deleted / Chat Not Found: {deleted}\n• ⚠️ Other Errors: {other}",
        "admin_rej_custom_prompt": "📝 **Type the custom rejection reason:**",
        "btn_admin_menu": "🟢 Admin Menu",
        "admin_bulk_app_prompt": "✅ **Bulk Approve:**\n\nPaste 50-200 Usernames / UIDs / Gmails here (one per line or separated by space/comma):\n\nBot will match and **Approve ONLY the matched tasks**.",
        "admin_bulk_rej_prompt": "❌ **Bulk Reject:**\n\nPaste 50-200 Usernames / UIDs / Gmails here (one per line or separated by space/comma):\n\nBot will match and **Reject ONLY the matched tasks**.",
    }
}

def get_user_lang(user_id: int) -> str:
    if not user_id:
        return 'bn'
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT language FROM users WHERE user_id = ?", (user_id,))
            row = cursor.fetchone()
            if row and row["language"]:
                return row["language"]
    except Exception:
        pass
    return 'bn'

def tr(user_id_or_lang: Union[int, str, None], key: str, **kwargs) -> str:
    if isinstance(user_id_or_lang, int):
        lang = get_user_lang(user_id_or_lang)
    elif isinstance(user_id_or_lang, str):
        lang = user_id_or_lang if user_id_or_lang in ('bn', 'en') else 'bn'
    else:
        lang = 'bn'

    lang_dict = MESSAGES.get(lang, MESSAGES['bn'])
    msg_template = lang_dict.get(key, MESSAGES['bn'].get(key, ''))
    if kwargs and msg_template:
        try:
            return msg_template.format(**kwargs)
        except Exception:
            return msg_template
    return msg_template

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------------------------------------
def get_setting_val(key: str, default: str = "") -> str:
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM admin_settings WHERE key = ?", (key,))
            row = cursor.fetchone()
            return row["value"] if row else default
    except Exception:
        return default

def set_setting_val(key: str, value: str):
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO admin_settings (key, value) VALUES (?, ?)", (key, str(value)))
            conn.commit()
    except Exception as e:
        logger.error(f"Error setting setting val {key}: {e}")

def is_owner(user_id: int) -> bool:
    admin_id_setting = get_setting_val("admin_id", "8001997389")
    primary_owner = int(admin_id_setting.split(",")[0].strip()) if admin_id_setting else 8001997389
    return user_id == primary_owner

def is_admin(user_id: int) -> bool:
    if is_owner(user_id):
        return True
    
    admin_id_setting = get_setting_val("admin_id", "8001997389")
    admins = [int(x.strip()) for x in admin_id_setting.split(",") if x.strip().isdigit()]
    if user_id in admins:
        return True

    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id FROM sub_admins WHERE user_id = ?", (user_id,))
            if cursor.fetchone():
                return True
    except Exception:
        pass
    return False

def is_banned(user_id: int) -> bool:
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT is_banned FROM users WHERE user_id = ?", (user_id,))
            row = cursor.fetchone()
            return bool(row["is_banned"]) if row else False
    except Exception:
        return False

def get_user_data(user_id: int):
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
            return cursor.fetchone()
    except Exception:
        None

def register_user(tg_user, referred_by_id=None):
    user_id = tg_user.id
    if get_user_data(user_id):
        return

    ref_code = f"NX{secrets.token_hex(3).upper()}"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ref_by = None
    if referred_by_id and referred_by_id != user_id:
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT user_id FROM users WHERE user_id = ?", (referred_by_id,))
                if cursor.fetchone():
                    ref_by = referred_by_id
                    cursor.execute("UPDATE users SET ref_count = ref_count + 1 WHERE user_id = ?", (ref_by,))
                    conn.commit()
        except Exception as e:
            logger.error(f"Error updating ref count: {e}")

    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO users (user_id, username, full_name, balance, referral_code, referred_by, joined_at, language)
                VALUES (?, ?, ?, 0.0, ?, ?, ?, 'bn')
                """,
                (user_id, tg_user.username, tg_user.full_name, ref_code, ref_by, now)
            )
            conn.commit()
    except Exception as e:
        logger.error(f"Error registering user: {e}")

# MULTI-CHANNEL FORCE JOIN CHECKER
# -----------------------------------------------------------------------------
# Supported admin input formats:
#   @publicchannel
#   @publicchannel|https://t.me/publicchannel
#   --1004397984847|https://t.me/+invite_hash
# IMPORTANT: for private/invite-only channels the left side MUST be the real
# Telegram chat ID (usually -100xxxxxxxxxx) and the bot must be an admin there.
# -----------------------------------------------------------------------------
def parse_force_join_channels(raw_value: str) -> List[Tuple[str, str]]:
    """Parse and normalize the admin-configured force-join channel list."""
    channels: List[Tuple[str, str]] = []
    seen = set()
    raw_value = str(raw_value or "")

    for raw_line in re.split(r"[\r\n]+", raw_value):
        line = raw_line.strip()
        if not line:
            continue

        if "|" in line:
            identifier, join_link = line.split("|", 1)
            identifier = identifier.strip()
            join_link = join_link.strip()
        else:
            identifier = line
            join_link = ""

        # Accept a public t.me URL as the identifier when no explicit ID is given.
        if (identifier.startswith("https://t.me/") or identifier.startswith("http://t.me/")):
            tail = identifier.split("t.me/", 1)[1].strip("/")
            if tail and not tail.startswith("+") and not tail.startswith("joinchat/"):
                identifier = "@" + tail.lstrip("@")

        if not identifier:
            continue

        # Normalize @username but preserve numeric Telegram chat IDs.
        if not identifier.startswith("-") and not identifier.isdigit() and not identifier.startswith("@"):
            identifier = "@" + identifier

        # If no explicit join URL is supplied, generate a public-channel URL.
        if not join_link:
            if identifier.startswith("@"):
                join_link = f"https://t.me/{identifier[1:]}"
            elif identifier.startswith("-") or identifier.isdigit():
                # A numeric private-channel ID cannot be converted into a join URL.
                # Admin must supply the actual invite link after '|'.
                join_link = "https://t.me/"

        # Only allow safe Telegram links in the button.
        if not re.match(r"^https?://t\.me/", join_link, re.IGNORECASE):
            if identifier.startswith("@"):
                join_link = f"https://t.me/{identifier[1:]}"
            else:
                join_link = "https://t.me/"

        key = identifier.lower()
        if key not in seen:
            seen.add(key)
            channels.append((identifier, join_link))

    return channels


def _member_is_joined(member) -> bool:
    """Return True for every Telegram membership state that counts as joined."""
    status = getattr(member, "status", "")
    if status in {"creator", "administrator", "member"}:
        return True
    # Telegram can return 'restricted' while the user is still a member.
    if status == "restricted" and bool(getattr(member, "is_member", False)):
        return True
    return False


async def check_force_join(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> Tuple[bool, List[Tuple[str, str]]]:
    if is_admin(user_id):
        return True, []

    enabled = str(get_setting_val("force_join_enabled", "OFF")).strip().upper()
    if enabled != "ON":
        return True, []

    raw_channels = get_setting_val(
        "force_channels_list",
        "@TrustVaultMailsOfficial|https://t.me/TrustVaultMailsOfficial"
    )
    channels = parse_force_join_channels(raw_channels)

    if not channels:
        # Never lock every user out because the admin accidentally saved an empty list.
        logger.warning("Force Join is ON but force_channels_list is empty; allowing access.")
        return True, []

    unjoined: List[Tuple[str, str]] = []
    for channel_identifier, join_link in channels:
        try:
            chat_identifier = (
                int(channel_identifier)
                if channel_identifier.lstrip("-").isdigit()
                else channel_identifier
            )
            member = await context.bot.get_chat_member(
                chat_id=chat_identifier,
                user_id=user_id
            )
            if not _member_is_joined(member):
                unjoined.append((channel_identifier, join_link))
        except Forbidden as e:
            # Usually means the bot cannot inspect membership (e.g. it is not an
            # admin/member of the channel). Fail closed for security and log clearly.
            logger.error(
                "Force Join cannot inspect %s. Make the bot an administrator in that channel. Error: %s",
                channel_identifier, e
            )
            unjoined.append((channel_identifier, join_link))
        except BadRequest as e:
            logger.error(
                "Force Join BadRequest for %s. Check the chat ID/username and bot permissions. Error: %s",
                channel_identifier, e
            )
            unjoined.append((channel_identifier, join_link))
        except TelegramError as e:
            logger.error("Force Join Telegram error for %s: %s", channel_identifier, e)
            unjoined.append((channel_identifier, join_link))
        except Exception as e:
            logger.exception("Unexpected Force Join error for %s: %s", channel_identifier, e)
            unjoined.append((channel_identifier, join_link))

    return (len(unjoined) == 0), unjoined


async def send_force_join_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                 unjoined_channels: Optional[List[Tuple[str, str]]] = None):
    """Render the force-join buttons consistently from every entry point."""
    user = update.effective_user
    if not user:
        return
    user_id = user.id

    if unjoined_channels is None:
        _, unjoined_channels = await check_force_join(user_id, context)

    buttons = []
    for idx, (_channel_id, link) in enumerate(unjoined_channels, 1):
        if link == "https://t.me/":
            # Do not render a useless URL button for a numeric ID without an invite.
            continue
        buttons.append([
            InlineKeyboardButton(f"🔗 {tr(user_id, 'btn_join_channel')} {idx}", url=link)
        ])

    buttons.append([
        InlineKeyboardButton(tr(user_id, "btn_verify"), callback_data="check_force_join_cb")
    ])

    text = tr(user_id, "force_join_msg")
    if not buttons[:-1]:
        text += "\n\n⚠️ Admin has not configured valid join links yet."

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.MARKDOWN
        )
    elif update.message:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.MARKDOWN
        )


def generate_random_username() -> str:
    prefix = "ig_"
    chars = "abcdefghijklmnopqrstuvwxyz0123456789"
    target_len = random.choice([13, 14])
    random_str = ''.join(random.choices(chars, k=target_len - len(prefix)))
    return f"{prefix}{random_str}"

def is_menu_or_admin_button(text: str) -> bool:
    if not text:
        return False
    known_buttons = [
        "💼 কাজ (Task)", "💼 Task", "👛 ওয়ালেট (Wallet)", "👛 Wallet",
        "👥 আমার রেফারেল", "👥 My Referrals", "🏆 লিডারবোর্ড", "🏆 Leaderboard",
        "💸 টাকা উত্তোলন (Withdraw)", "💸 Withdraw", "🎧 সাপোর্ট", "🎧 Support",
        "🌐 ভাষা পরিবর্তন", "🌐 Language", "🔑 Admin Panel", "🟢 Admin Panel", "🔙 মূল মেনু", "🟢 মূল মেনু", "🔙 Main Menu", "🟢 Main Menu",
        "❌ বাতিল", "❌ Cancel", "📸 ইনস্টাগ্রাম", "📸 Instagram", "📘 ফেসবুক", "📘 Facebook",
        "✉️ জিমেইল", "✉️ Gmail", "📩 পেন্ডিং টাস্ক", "🟢 📩 পেন্ডিং টাস্ক", "📩 Pending Tasks", "🟢 📩 Pending Tasks",
        "⌛ হোল্ড টাস্ক", "🟢 ⌛ হোল্ড টাস্ক", "⌛ Held Tasks", "🟢 ⌛ Held Tasks",
        "✅ Approve All (বাল্ক এপ্রুভ)", "🟢 ✅ বাল্ক এপ্রুভ", "✅ Approve All", "🟢 ✅ Approve All",
        "❌ Reject All (বাল্ক রিজেক্ট)", "🟢 ❌ বাল্ক রিজেক্ট", "❌ Reject All", "🟢 ❌ Reject All",
        "🔑 পাসওয়ার্ড ম্যানেজ", "🟢 🔑 পাসওয়ার্ড ম্যানেজ", "🔑 Password Manager", "🟢 🔑 Password Manager",
        "🏷️ প্রাইস ম্যানেজ", "🟢 🏷️ প্রাইস ম্যানেজ", "🏷️ Price Manager", "🟢 🏷️ Price Manager",
        "🔍 টাস্ক সার্চ", "🟢 🔍 টাস্ক সার্চ", "🔍 Task Search", "🟢 🔍 Task Search",
        "📂 ইউজার এক্সপোর্ট", "🟢 📂 ইউজার এক্সপোর্ট", "📂 User Export", "🟢 📂 User Export",
        "📊 এক্সেল শীট (Spreadsheet)", "🟢 📊 এক্সেল শীট (Spreadsheet)", "📊 Spreadsheet UI", "🟢 📊 Spreadsheet UI",
        "💸 উইথড্র রিকোয়েস্ট", "🟢 💸 উইথড্র রিকোয়েস্ট", "💸 Withdraw Requests", "🟢 💸 Withdraw Requests",
        "👥 ইউজার ম্যানেজমেন্ট", "🟢 👥 ইউজার ম্যানেজমেন্ট", "👥 User Management", "🟢 👥 User Management",
        "👑 এডমিন কন্ট্রোল", "🟢 👑 এডমিন কন্ট্রোল", "👑 Admin Control", "🟢 👑 Admin Control",
        "⚙️ সিস্টেম সেটিংস", "🟢 ⚙️ সিস্টেম সেটিংস", "⚙️ System Settings", "🟢 ⚙️ System Settings",
        "📊 লাইভ স্ট্যাটাস", "🟢 📊 লাইভ স্ট্যাটাস", "📊 Live Stats", "🟢 📊 Live Stats",
        "📢 ব্রডকাস্ট", "🟢 📢 ব্রডকাস্ট", "📢 Broadcast", "🟢 📢 Broadcast"
    ]
    return text in known_buttons or "বাল্ক এপ্রুভ" in text or "Bulk Approve" in text or "বাল্ক রিজেক্ট" in text or "Bulk Reject" in text

FIRST_NAMES = ["Sohan", "Kamal", "Robi", "Arif", "Nayan", "Sumon", "Rashed", "Alim", "Nusrat", "Riya", "Fariha", "Mim"]
LAST_NAMES = ["Ahmed", "Sarker", "Hasan", "Rahman", "Uddin", "Chowdhury", "Akter", "Khatun", "Islam", "Khan"]

# -----------------------------------------------------------------------------
# CONVERSATION STATES
# -----------------------------------------------------------------------------
(
    STATE_TASK_CATEGORY,
    STATE_TASK_INSTAGRAM,
    STATE_TASK_FACEBOOK,
    STATE_TASK_IG_2FA_INPUT,
    STATE_TASK_IG_SEED_INPUT,
    STATE_TASK_FB_2FA_INPUT,
    STATE_TASK_FB_COOKIES_UID,
    STATE_TASK_FB_COOKIES_STR,
    STATE_TASK_GMAIL_INPUT,
    STATE_TASK_2FA_SECRET,
    STATE_TASK_FINAL_SUBMIT,
    STATE_WD_METHOD,
    STATE_WD_NUMBER,
    STATE_WD_AMOUNT,
    STATE_ADMIN_SEARCH_QUERY,
    STATE_ADMIN_EXPORT_USER,
    STATE_ADMIN_BROADCAST,
    STATE_ADMIN_REJECT_REASON,
    STATE_ADMIN_SETTING_VAL,
    STATE_ADMIN_BULK_APPROVE,
    STATE_ADMIN_BULK_REJECT,
    STATE_ADMIN_USER_MGMT,
    STATE_ADMIN_ADD_SUBADMIN,
    STATE_ADMIN_REM_SUBADMIN,
    STATE_ADMIN_BAN_USER,
    STATE_ADMIN_UNBAN_USER,
    STATE_ADMIN_ADD_BAL,
    STATE_ADMIN_SUB_BAL,
    STATE_ADMIN_USER_SEARCH,
    STATE_ADMIN_FORCE_CHANNELS_EDIT,
) = range(30)

# -----------------------------------------------------------------------------
# DYNAMIC MULTILINGUAL REPLY KEYBOARDS
# -----------------------------------------------------------------------------

def get_main_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    if user_id and is_admin(user_id):
        keyboard = [
            [KeyboardButton(tr(l, "btn_task")), KeyboardButton(tr(l, "btn_wallet"))],
            [KeyboardButton(tr(l, "btn_ref")), KeyboardButton(tr(l, "btn_leaderboard"))],
            [KeyboardButton(tr(l, "btn_withdraw")), KeyboardButton(tr(l, "btn_support"))],
            [KeyboardButton(tr(l, "btn_lang"))],
            [KeyboardButton(tr(l, "btn_admin_panel"))]
        ]
    else:
        keyboard = [
            [KeyboardButton(tr(l, "btn_task")), KeyboardButton(tr(l, "btn_wallet"))],
            [KeyboardButton(tr(l, "btn_ref")), KeyboardButton(tr(l, "btn_leaderboard"))],
            [KeyboardButton(tr(l, "btn_withdraw")), KeyboardButton(tr(l, "btn_support"))],
            [KeyboardButton(tr(l, "btn_lang"))]
        ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_task_category_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_instagram")), KeyboardButton(tr(l, "btn_facebook"))],
        [KeyboardButton(tr(l, "btn_gmail"))],
        [KeyboardButton(tr(l, "btn_main_menu"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_instagram_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_ig_2fa")), KeyboardButton(tr(l, "btn_ig_seed"))],
        [KeyboardButton(tr(l, "btn_main_menu")), KeyboardButton(tr(l, "btn_cancel"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_task_done_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_done"))],
        [KeyboardButton(tr(l, "btn_cancel")), KeyboardButton(tr(l, "btn_main_menu"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

get_seed_done_keyboard = get_task_done_keyboard

def get_instagram_task_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_gen_2fa"))],
        [KeyboardButton(tr(l, "btn_acct_done"))],
        [KeyboardButton(tr(l, "btn_cancel"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_facebook_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_fb_2fa")), KeyboardButton(tr(l, "btn_fb_cookies"))],
        [KeyboardButton(tr(l, "btn_main_menu")), KeyboardButton(tr(l, "btn_cancel"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_wallet_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_withdraw"))],
        [KeyboardButton(tr(l, "btn_main_menu"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_withdraw_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    bkash_active = get_setting_val("wd_bkash_active", "ON") == "ON"
    nagad_active = get_setting_val("wd_nagad_active", "ON") == "ON"
    usdt_active = get_setting_val("wd_usdt_active", "ON") == "ON"

    keyboard = []
    mfs_row = []
    if bkash_active:
        mfs_row.append(KeyboardButton(tr(l, "btn_bkash")))
    if nagad_active:
        mfs_row.append(KeyboardButton(tr(l, "btn_nagad")))
    if mfs_row:
        keyboard.append(mfs_row)

    if usdt_active:
        keyboard.append([KeyboardButton(tr(l, "btn_usdt"))])

    keyboard.append([KeyboardButton(tr(l, "btn_main_menu"))])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_cancel_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_cancel"))],
        [KeyboardButton(tr(l, "btn_main_menu"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_admin_keyboard(user_id: int = None, lang: str = None) -> ReplyKeyboardMarkup:
    l = lang or get_user_lang(user_id)
    keyboard = [
        [KeyboardButton(tr(l, "btn_pending_tasks")), KeyboardButton(tr(l, "btn_held_tasks"))],
        [KeyboardButton(tr(l, "btn_bulk_approve")), KeyboardButton(tr(l, "btn_bulk_reject"))],
        [KeyboardButton(tr(l, "btn_pwd_mgr")), KeyboardButton(tr(l, "btn_price_mgr"))],
        [KeyboardButton(tr(l, "btn_task_search")), KeyboardButton(tr(l, "btn_user_export"))],
        [KeyboardButton(tr(l, "btn_excel_export")), KeyboardButton(tr(l, "btn_wd_reqs"))],
        [KeyboardButton(tr(l, "btn_user_mgmt")), KeyboardButton(tr(l, "btn_admin_mgmt"))],
        [KeyboardButton(tr(l, "btn_sys_settings")), KeyboardButton(tr(l, "btn_live_stats"))],
        [KeyboardButton(tr(l, "btn_broadcast")), KeyboardButton(tr(l, "btn_main_menu"))]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def clear_user_state(context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()

async def safe_delete_message(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_id: int):
    """Safely delete message without throwing errors"""
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
    except Exception:
        pass

# -----------------------------------------------------------------------------
# START COMMAND & FORCE JOIN
# -----------------------------------------------------------------------------

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user = update.effective_user
    user_id = user.id

    if get_setting_val("maintenance_mode") == "ON" and not is_admin(user_id):
        await update.message.reply_text(
            tr(user_id, "maintenance_msg"),
            parse_mode=ParseMode.MARKDOWN
        )
        return ConversationHandler.END

    referred_by_id = None
    if context.args and len(context.args) > 0 and context.args[0].isdigit():
        referred_by_id = int(context.args[0])

    is_new_user = (get_user_data(user_id) is None)
    register_user(user, referred_by_id)

    # INSTANT REFERRAL NOTIFICATION TO REFERRER
    if is_new_user and referred_by_id and referred_by_id != user_id:
        try:
            ref_user = get_user_data(referred_by_id)
            if ref_user:
                await context.bot.send_message(
                    chat_id=referred_by_id,
                    text=tr(referred_by_id, "ref_new_user_joined", name=html.escape(user.first_name), user_id=user_id),
                    parse_mode=ParseMode.HTML
                )
        except Exception as e:
            logger.error(f"Failed to send referral alert to {referred_by_id}: {e}")

    if is_banned(user_id):
        await update.message.reply_text(tr(user_id, "banned_msg"))
        return ConversationHandler.END

    is_joined, unjoined_channels = await check_force_join(user_id, context)
    if not is_joined:
        await send_force_join_prompt(update, context, unjoined_channels)
        return ConversationHandler.END

    welcome_msg = tr(user_id, "welcome_msg", name=html.escape(user.first_name))

    await update.message.reply_text(
        welcome_msg,
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return ConversationHandler.END

async def check_force_join_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id

    is_joined, unjoined_channels = await check_force_join(user_id, context)
    if is_joined:
        await query.answer(tr(user_id, "force_join_success"), show_alert=True)
        await safe_delete_message(context, query.message.chat_id, query.message.message_id)
        await context.bot.send_message(
            chat_id=user_id,
            text=tr(user_id, "force_join_success"),
            reply_markup=get_main_keyboard(user_id),
            parse_mode=ParseMode.MARKDOWN
        )
    else:
        await query.answer(tr(user_id, "force_join_fail"), show_alert=True)
        # Keep the verification UI visible so the user can immediately retry.
        try:
            await send_force_join_prompt(update, context, unjoined_channels)
        except Exception as e:
            logger.error("Failed to refresh Force Join prompt: %s", e)

# -----------------------------------------------------------------------------
# TASK WORKFLOWS
# -----------------------------------------------------------------------------

async def handle_task_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    if is_banned(user_id):
        await update.message.reply_text(tr(user_id, "banned_msg"))
        return ConversationHandler.END

    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    await update.message.reply_text(
        tr(user_id, "task_cat_prompt"),
        reply_markup=get_task_category_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_TASK_CATEGORY

async def handle_instagram_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if is_banned(user_id):
        return ConversationHandler.END

    ig_2fa_active = get_setting_val("ig_2fa_active", "ON")
    ig_seed_active = get_setting_val("ig_seed_active", "ON")

    if ig_2fa_active == "OFF" and ig_seed_active == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_ig"), reply_markup=get_task_category_keyboard(user_id))
        return STATE_TASK_CATEGORY

    await update.message.reply_text(
        tr(user_id, "ig_menu_title"),
        reply_markup=get_instagram_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_TASK_INSTAGRAM

async def handle_facebook_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if is_banned(user_id):
        return ConversationHandler.END

    fb_2fa_active = get_setting_val("fb_2fa_active", "ON")
    fb_cookies_active = get_setting_val("fb_cookies_active", "ON")

    if fb_2fa_active == "OFF" and fb_cookies_active == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_fb"), reply_markup=get_task_category_keyboard(user_id))
        return STATE_TASK_CATEGORY

    await update.message.reply_text(
        tr(user_id, "fb_menu_title"),
        reply_markup=get_facebook_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_TASK_FACEBOOK

async def handle_gmail_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if is_banned(user_id):
        return ConversationHandler.END

    if get_setting_val("gmail_task_active", "ON") == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_gmail"), reply_markup=get_task_category_keyboard(user_id))
        return STATE_TASK_CATEGORY

    price = get_setting_val("gmail_task_price", "22.00")
    fixed_pass = get_setting_val("gmail_default_password", "aass1122")

    context.user_data["task_type"] = "gmail"
    context.user_data["reward_amount"] = float(price)
    context.user_data["assigned_pass"] = fixed_pass

    msg = tr(user_id, "gmail_guidelines", price=price, fixed_pass=fixed_pass)
    sent_m = await update.message.reply_text(msg, reply_markup=get_task_done_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["cred_msg_id"] = sent_m.message_id
    return STATE_TASK_GMAIL_INPUT

# --- INSTAGRAM WORKFLOWS ---
async def handle_ig_2fa_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if get_setting_val("ig_2fa_active", "ON") == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_ig_2fa"), reply_markup=get_instagram_keyboard(user_id))
        return STATE_TASK_INSTAGRAM

    random_username = generate_random_username()
    fixed_pass = get_setting_val("ig_default_password", "RBKpass@06")
    price = get_setting_val("ig_2fa_price", "4.00")

    context.user_data["task_type"] = "ig_2fa"
    context.user_data["reward_amount"] = float(price)
    context.user_data["assigned_username"] = random_username
    context.user_data["assigned_pass"] = fixed_pass
    context.user_data["waiting_for_secret"] = True

    msg = tr(user_id, "ig_2fa_guidelines", price=price, username=random_username, fixed_pass=fixed_pass)
    sent_m = await update.message.reply_text(msg, reply_markup=get_instagram_task_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["cred_msg_id"] = sent_m.message_id
    return STATE_TASK_IG_2FA_INPUT

async def handle_ig_seed_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if get_setting_val("ig_seed_active", "ON") == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_ig_seed"), reply_markup=get_instagram_keyboard(user_id))
        return STATE_TASK_INSTAGRAM

    price = get_setting_val("ig_seed_price", "4.30")
    context.user_data["task_type"] = "ig_seed"
    context.user_data["reward_amount"] = float(price)

    msg = tr(user_id, "ig_seed_guidelines", price=price)
    sent_m = await update.message.reply_text(msg, reply_markup=get_task_done_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["cred_msg_id"] = sent_m.message_id
    return STATE_TASK_IG_SEED_INPUT

# --- FACEBOOK WORKFLOWS ---
async def handle_fb_2fa_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if get_setting_val("fb_2fa_active", "ON") == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_fb_2fa"), reply_markup=get_facebook_keyboard(user_id))
        return STATE_TASK_FACEBOOK

    first_name = random.choice(FIRST_NAMES)
    last_name = random.choice(LAST_NAMES)
    fixed_pass = get_setting_val("fb_default_password", "FBKpass@07")
    price = get_setting_val("fb_2fa_price", "5.00")

    context.user_data["task_type"] = "fb_2fa"
    context.user_data["reward_amount"] = float(price)
    context.user_data["fb_first"] = first_name
    context.user_data["fb_last"] = last_name
    context.user_data["assigned_pass"] = fixed_pass
    context.user_data["waiting_for_secret"] = True

    msg = tr(user_id, "fb_2fa_guidelines", price=price, first_name=first_name, last_name=last_name, fixed_pass=fixed_pass)
    sent_m = await update.message.reply_text(msg, reply_markup=get_instagram_task_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["cred_msg_id"] = sent_m.message_id
    return STATE_TASK_FB_2FA_INPUT

async def handle_fb_cookies_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if get_setting_val("fb_cookies_active", "ON") == "OFF":
        await update.message.reply_text(tr(user_id, "task_off_fb_cookies"), reply_markup=get_facebook_keyboard(user_id))
        return STATE_TASK_FACEBOOK

    first_name = random.choice(FIRST_NAMES)
    last_name = random.choice(LAST_NAMES)
    fixed_pass = get_setting_val("fb_default_password", "FBKpass@07")
    price = get_setting_val("fb_cookies_price", "7.00")

    context.user_data["task_type"] = "fb_cookies"
    context.user_data["reward_amount"] = float(price)
    context.user_data["fb_first"] = first_name
    context.user_data["fb_last"] = last_name
    context.user_data["assigned_pass"] = fixed_pass

    msg = tr(user_id, "fb_cookies_guidelines_uid", price=price, first_name=first_name, last_name=last_name, fixed_pass=fixed_pass)
    sent_m = await update.message.reply_text(msg, reply_markup=get_task_done_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["cred_msg_id"] = sent_m.message_id
    return STATE_TASK_FB_COOKIES_UID

# --- 2FA OTP VERIFICATION PROCESS ---
async def process_2fa_secret(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if is_menu_or_admin_button(text):
        clear_user_state(context)
        return ConversationHandler.END

    secret_input = text.replace(" ", "").upper()

    try:
        totp = pyotp.TOTP(secret_input)
        current_otp = totp.now()
    except Exception:
        await update.message.reply_text(
            tr(user_id, "otp_invalid_secret"),
            parse_mode=ParseMode.MARKDOWN
        )
        return STATE_TASK_2FA_SECRET

    context.user_data["secret_key"] = secret_input
    context.user_data["current_otp"] = current_otp
    context.user_data["waiting_for_secret"] = False

    task_type = context.user_data.get("task_type", "unknown")
    msg = tr(user_id, "otp_verified_msg", secret=secret_input, otp=current_otp)
    if task_type == "ig_2fa":
        assigned = context.user_data.get("assigned_username")
        msg += f"\n\n(Username: <code>{assigned}</code>)"

    sent_m = await update.message.reply_text(msg, reply_markup=get_instagram_task_keyboard(user_id), parse_mode=ParseMode.HTML)
    context.user_data["otp_msg_id"] = sent_m.message_id
    return STATE_TASK_FINAL_SUBMIT

# --- FINAL SUBMISSION HANDLERS ---
async def process_task_finalize(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text

    if is_menu_or_admin_button(text) and text not in ["🔑 2FA Code Generate", "🔑 2FA Code Generate করুন", "🔑 Generate 2FA Code", "✅ একাউন্ট খোলা শেষ", "✅ Account Creation Done"]:
        clear_user_state(context)
        return ConversationHandler.END

    if text in ["🔑 2FA Code Generate", "🔑 2FA Code Generate করুন", "🔑 Generate 2FA Code"]:
        context.user_data["waiting_for_secret"] = True
        await update.message.reply_text(
            tr(user_id, "otp_prompt_secret"),
            reply_markup=get_cancel_keyboard(user_id)
        )
        return STATE_TASK_2FA_SECRET

    if text in ["✅ একাউন্ট খোলা শেষ", "✅ Account Creation Done"]:
        if context.user_data.get("waiting_for_secret", True):
            await update.message.reply_text(
                tr(user_id, "otp_need_verify_first"),
                reply_markup=get_instagram_task_keyboard(user_id)
            )
            return STATE_TASK_FINAL_SUBMIT

        # AUTO DELETE SECURITY CREDENTIAL MESSAGES FOR USER SAFETY
        if "cred_msg_id" in context.user_data:
            await safe_delete_message(context, update.effective_chat.id, context.user_data["cred_msg_id"])
        if "otp_msg_id" in context.user_data:
            await safe_delete_message(context, update.effective_chat.id, context.user_data["otp_msg_id"])

        task_type = context.user_data.get("task_type")
        password = context.user_data.get("assigned_pass")
        secret_key = context.user_data.get("secret_key")
        reward = context.user_data.get("reward_amount", 4.0)

        if task_type == "ig_2fa":
            username = context.user_data.get("assigned_username", generate_random_username())
        else:
            await update.message.reply_text(
                tr(user_id, "enter_username_prompt"),
                reply_markup=get_cancel_keyboard(user_id)
            )
            return STATE_TASK_FINAL_SUBMIT

        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO task_submissions 
                (user_id, task_type, submitted_username, submitted_password, secret_key_2fa, status, reward_amount, is_archived, created_at, updated_at) 
                VALUES (?, ?, ?, ?, ?, 'Pending', ?, 0, ?, ?)
                """,
                (user_id, task_type, username, password, secret_key, reward, now, now)
            )
            submission_id = cursor.lastrowid
            conn.commit()

        task_display = "Instagram 2FA" if task_type == "ig_2fa" else "Facebook 2FA"
        await update.message.reply_text(
            tr(user_id, "task_submitted_success", sub_id=submission_id, task_display=task_display, username=username),
            reply_markup=get_main_keyboard(user_id),
            parse_mode=ParseMode.HTML
        )
        await notify_admins_new_submission(context, submission_id, user_id, task_type, username, password, secret_key, "", reward)
        clear_user_state(context)
        return ConversationHandler.END

    username = text.strip()
    task_type = context.user_data.get("task_type")
    password = context.user_data.get("assigned_pass")
    secret_key = context.user_data.get("secret_key")
    reward = context.user_data.get("reward_amount", 5.0)

    if "cred_msg_id" in context.user_data:
        await safe_delete_message(context, update.effective_chat.id, context.user_data["cred_msg_id"])
    if "otp_msg_id" in context.user_data:
        await safe_delete_message(context, update.effective_chat.id, context.user_data["otp_msg_id"])

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO task_submissions 
            (user_id, task_type, submitted_username, submitted_password, secret_key_2fa, status, reward_amount, is_archived, created_at, updated_at) 
            VALUES (?, ?, ?, ?, ?, 'Pending', ?, 0, ?, ?)
            """,
            (user_id, task_type, username, password, secret_key, reward, now, now)
        )
        submission_id = cursor.lastrowid
        conn.commit()

    task_display = "Instagram 2FA" if task_type == "ig_2fa" else "Facebook 2FA"
    await update.message.reply_text(
        tr(user_id, "task_submitted_success", sub_id=submission_id, task_display=task_display, username=username),
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.HTML
    )
    await notify_admins_new_submission(context, submission_id, user_id, task_type, username, password, secret_key, "", reward)
    clear_user_state(context)
    return ConversationHandler.END

async def handle_ig_seed_file_or_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if update.message.document:
        doc = update.message.document
        context.user_data["seed_file_id"] = doc.file_id
        context.user_data["seed_file_name"] = doc.file_name or "spreadsheet.xlsx"
        context.user_data["has_seed_data"] = True
    elif update.message.text:
        text = update.message.text.strip()
        if is_menu_or_admin_button(text):
            clear_user_state(context)
            return ConversationHandler.END
        
        context.user_data["seed_text"] = text
        context.user_data["has_seed_data"] = True

    await update.message.reply_text(
        tr(user_id, "file_received_msg"),
        reply_markup=get_seed_done_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_TASK_IG_SEED_INPUT

async def process_ig_seed_finalize(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if not context.user_data.get("has_seed_data"):
        await update.message.reply_text(
            tr(user_id, "no_file_or_data_err"),
            reply_markup=get_seed_done_keyboard(user_id),
            parse_mode=ParseMode.MARKDOWN
        )
        return STATE_TASK_IG_SEED_INPUT

    if "cred_msg_id" in context.user_data:
        await safe_delete_message(context, update.effective_chat.id, context.user_data["cred_msg_id"])

    reward = context.user_data.get("reward_amount", 4.30)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    file_id = context.user_data.get("seed_file_id")
    file_name = context.user_data.get("seed_file_name", "Seed_Data")
    seed_text = context.user_data.get("seed_text", "")

    if file_id:
        submitted_name = f"FILE: {file_name}"
        cookies_val = f"FILE_ID:{file_id}"
    else:
        submitted_name = seed_text[:100]
        cookies_val = seed_text

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO task_submissions 
            (user_id, task_type, submitted_username, submitted_password, secret_key_2fa, cookies_data, status, reward_amount, is_archived, created_at, updated_at) 
            VALUES (?, 'ig_seed', ?, 'N/A', 'N/A', ?, 'Pending', ?, 0, ?, ?)
            """,
            (user_id, submitted_name, cookies_val, reward, now, now)
        )
        submission_id = cursor.lastrowid
        conn.commit()

    await update.message.reply_text(
        tr(user_id, "ig_seed_submitted_success", sub_id=submission_id),
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )

    await notify_admins_new_submission(
        context, submission_id, user_id, "ig_seed", submitted_name, "N/A", "N/A", cookies_val, reward, file_id=file_id
    )
    clear_user_state(context)
    return ConversationHandler.END

async def process_fb_cookies_uid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    uid = update.message.text.strip()

    if is_menu_or_admin_button(uid):
        clear_user_state(context)
        return ConversationHandler.END

    context.user_data["fb_uid"] = uid
    await update.message.reply_text(
        tr(user_id, "fb_cookies_guidelines_str"),
        reply_markup=get_task_done_keyboard(user_id)
    )
    return STATE_TASK_FB_COOKIES_STR

async def process_fb_cookies_str(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    cookie_str = update.message.text.strip()

    if is_menu_or_admin_button(cookie_str):
        clear_user_state(context)
        return ConversationHandler.END

    if "cred_msg_id" in context.user_data:
        await safe_delete_message(context, update.effective_chat.id, context.user_data["cred_msg_id"])

    uid = context.user_data.get("fb_uid")
    password = context.user_data.get("assigned_pass")
    reward = context.user_data.get("reward_amount", 7.0)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO task_submissions 
            (user_id, task_type, submitted_username, submitted_password, secret_key_2fa, cookies_data, status, reward_amount, is_archived, created_at, updated_at) 
            VALUES (?, 'fb_cookies', ?, ?, 'N/A', ?, 'Pending', ?, 0, ?, ?)
            """,
            (user_id, uid, password, cookie_str, reward, now, now)
        )
        submission_id = cursor.lastrowid
        conn.commit()

    await update.message.reply_text(
        tr(user_id, "fb_cookies_submitted_success", sub_id=submission_id),
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    await notify_admins_new_submission(context, submission_id, user_id, "fb_cookies", uid, password, "N/A", cookie_str, reward)
    clear_user_state(context)
    return ConversationHandler.END

async def process_gmail_submission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    gmail_acc = update.message.text.strip()

    if is_menu_or_admin_button(gmail_acc):
        clear_user_state(context)
        return ConversationHandler.END

    if "cred_msg_id" in context.user_data:
        await safe_delete_message(context, update.effective_chat.id, context.user_data["cred_msg_id"])

    password = context.user_data.get("assigned_pass", "aass1122")
    reward = context.user_data.get("reward_amount", 22.0)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO task_submissions 
            (user_id, task_type, submitted_username, submitted_password, secret_key_2fa, status, reward_amount, is_archived, created_at, updated_at) 
            VALUES (?, 'gmail', ?, ?, 'N/A', 'Pending', ?, 0, ?, ?)
            """,
            (user_id, gmail_acc, password, reward, now, now)
        )
        submission_id = cursor.lastrowid
        conn.commit()

    await update.message.reply_text(
        tr(user_id, "gmail_submitted_success", sub_id=submission_id),
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    await notify_admins_new_submission(context, submission_id, user_id, "gmail", gmail_acc, password, "N/A", "", reward)
    clear_user_state(context)
    return ConversationHandler.END

# --- ADMIN NOTIFICATION HELPER ---
async def notify_admins_new_submission(context, sub_id, user_id, task_type, username, password, secret_key, cookies, reward, file_id=None):
    admin_id_setting = get_setting_val("admin_id", "8001997389")
    admins = [int(x.strip()) for x in admin_id_setting.split(",") if x.strip().isdigit()]

    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id FROM sub_admins")
            for row in cursor.fetchall():
                admins.append(row["user_id"])
    except Exception:
        pass
    admins = list(set(admins))

    task_icon = "📸" if "ig" in task_type else "📘" if "fb" in task_type else "✉️"
    task_name = {
        "ig_2fa": "Instagram 2FA",
        "ig_seed": "Instagram Seed",
        "fb_2fa": "Facebook 2FA",
        "fb_cookies": "Facebook Cookies",
        "gmail": "Gmail"
    }.get(task_type, task_type.upper())

    admin_msg = (
        f"📥 **নতুন টাস্ক সাবমিট হয়েছে! / New Task Submitted!**\n\n"
        f"🆔 **টাস্ক আইডি / Task ID:** #{sub_id}\n"
        f"👤 **ইউজার আইডি / User ID:** <code>{user_id}</code>\n"
        f"📱 **টাইপ / Type:** {task_icon} {task_name}\n"
        f"👤 **সাবমিশন / Submission:** <code>{username}</code>\n"
        f"🔑 **পাসওয়ার্ড / Password:** <code>{password}</code>\n"
    )
    if secret_key and secret_key != "N/A":
        admin_msg += f"🔐 **2FA Secret:** <code>{secret_key}</code>\n"
    if cookies:
        admin_msg += f"🍪 **Cookies:** <code>{cookies[:50]}...</code>\n"
    admin_msg += f"💵 **মূল্য / Price:** ৳{reward:.2f}"

    buttons = [
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"adm_app_sub_{sub_id}"),
            InlineKeyboardButton("⏳ Hold", callback_data=f"adm_hold_sub_{sub_id}")
        ],
        [InlineKeyboardButton("❌ Reject", callback_data=f"adm_rej_prompt_{sub_id}")]
    ]

    for admin in admins:
        try:
            if file_id:
                await context.bot.send_document(
                    chat_id=admin,
                    document=file_id,
                    caption=admin_msg,
                    reply_markup=InlineKeyboardMarkup(buttons),
                    parse_mode=ParseMode.HTML
                )
            else:
                await context.bot.send_message(
                    chat_id=admin,
                    text=admin_msg,
                    reply_markup=InlineKeyboardMarkup(buttons),
                    parse_mode=ParseMode.HTML
                )
        except Exception as e:
            logger.error(f"Failed to notify admin {admin}: {e}")

# -----------------------------------------------------------------------------
# WALLET & WITHDRAWAL SYSTEM
# -----------------------------------------------------------------------------

async def handle_wallet_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    if is_banned(user_id):
        return ConversationHandler.END

    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    user_data = get_user_data(user_id)
    if not user_data:
        register_user(update.effective_user)
        user_data = get_user_data(user_id)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(amount) FROM withdrawals WHERE user_id = ? AND status = 'Pending'", (user_id,))
        p_row = cursor.fetchone()
        pending_withdraw = p_row[0] if p_row[0] else 0.0

        cursor.execute("SELECT SUM(reward_amount) FROM task_submissions WHERE user_id = ? AND status = 'Approved'", (user_id,))
        a_row = cursor.fetchone()
        approved_earnings = a_row[0] if a_row[0] else 0.0

        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE user_id = ? AND status = 'Approved'", (user_id,))
        completed_tasks = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE user_id = ? AND status = 'Pending'", (user_id,))
        pending_tasks = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE user_id = ? AND status = 'Hold'", (user_id,))
        held_tasks = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE user_id = ? AND status = 'Rejected'", (user_id,))
        rejected_tasks = cursor.fetchone()[0]

    lifetime_earnings = approved_earnings + user_data["ref_earnings"]

    wallet_text = tr(
        user_id, "wallet_text",
        balance=user_data['balance'],
        pending_wd=pending_withdraw,
        lifetime=lifetime_earnings,
        ref_earnings=user_data['ref_earnings'],
        completed=completed_tasks,
        pending=pending_tasks,
        held=held_tasks,
        rejected=rejected_tasks
    )

    await update.message.reply_text(
        wallet_text,
        reply_markup=get_wallet_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return ConversationHandler.END

async def handle_withdraw_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    if is_banned(user_id):
        return ConversationHandler.END

    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    bkash_active = get_setting_val("wd_bkash_active", "ON") == "ON"
    nagad_active = get_setting_val("wd_nagad_active", "ON") == "ON"
    usdt_active = get_setting_val("wd_usdt_active", "ON") == "ON"

    if not bkash_active and not nagad_active and not usdt_active:
        await update.message.reply_text(tr(user_id, "wd_all_disabled"), reply_markup=get_main_keyboard(user_id))
        return ConversationHandler.END

    await update.message.reply_text(
        tr(user_id, "wd_gateway_prompt"),
        reply_markup=get_withdraw_keyboard(user_id)
    )
    return STATE_WD_METHOD

async def handle_withdraw_method(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = update.effective_user.id

    if text in ["🔙 মূল মেনু", "🟢 মূল মেনু", "🔙 Main Menu", "🟢 Main Menu"]:
        await update.message.reply_text(tr(user_id, "returning_main_menu"), reply_markup=get_main_keyboard(user_id))
        return ConversationHandler.END

    method = None
    if text in ["🟢 বিকাশ (Bkash)", "🟢 Bkash"]:
        if get_setting_val("wd_bkash_active", "ON") == "ON":
            method = "bkash"
    elif text in ["🟠 নগদ (Nagad)", "🟠 Nagad"]:
        if get_setting_val("wd_nagad_active", "ON") == "ON":
            method = "nagad"
    elif text in ["🔵 USDT (BEP-20)"]:
        if get_setting_val("wd_usdt_active", "ON") == "ON":
            method = "usdt"

    if not method:
        await update.message.reply_text(tr(user_id, "wd_invalid_method"), reply_markup=get_withdraw_keyboard(user_id))
        return STATE_WD_METHOD

    context.user_data["wd_method"] = method

    if method == "usdt":
        msg = tr(user_id, "wd_enter_account_usdt")
    else:
        msg = tr(user_id, "wd_enter_account_mfs", method=method.upper())

    await update.message.reply_text(msg, reply_markup=get_cancel_keyboard(user_id))
    return STATE_WD_NUMBER

async def handle_withdraw_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    number = update.message.text.strip()

    if number in ["❌ বাতিল", "❌ Cancel", "🔙 মূল মেনু", "🟢 মূল মেনু", "🔙 Main Menu", "🟢 Main Menu"]:
        await update.message.reply_text(tr(user_id, "operation_cancelled"), reply_markup=get_main_keyboard(user_id))
        return ConversationHandler.END

    method = context.user_data.get("wd_method")

    if method != "usdt" and not re.match(r"^[0-9]{11,14}$", number):
        await update.message.reply_text(tr(user_id, "wd_invalid_account"), reply_markup=get_cancel_keyboard(user_id))
        return STATE_WD_NUMBER

    context.user_data["wd_number"] = number
    user_data = get_user_data(user_id)

    min_limit = float(get_setting_val(f"min_withdraw_{method}", "50.0"))
    max_limit = float(get_setting_val(f"max_withdraw_{method}", "5000.0"))

    msg = tr(user_id, "wd_enter_amount", balance=user_data['balance'], min_limit=min_limit, max_limit=max_limit)

    await update.message.reply_text(msg, reply_markup=get_cancel_keyboard(user_id))
    return STATE_WD_AMOUNT

async def handle_withdraw_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    amount_str = update.message.text.strip()

    if amount_str in ["❌ বাতিল", "❌ Cancel", "🔙 মূল মেনু", "🟢 মূল মেনু", "🔙 Main Menu", "🟢 Main Menu"]:
        await update.message.reply_text(tr(user_id, "operation_cancelled"), reply_markup=get_main_keyboard(user_id))
        return ConversationHandler.END

    try:
        amount = float(amount_str)
    except ValueError:
        await update.message.reply_text(tr(user_id, "wd_invalid_amount_number"), reply_markup=get_cancel_keyboard(user_id))
        return STATE_WD_AMOUNT

    user_data = get_user_data(user_id)
    method = context.user_data.get("wd_method")

    min_limit = float(get_setting_val(f"min_withdraw_{method}", "50.0"))
    max_limit = float(get_setting_val(f"max_withdraw_{method}", "5000.0"))

    if amount < min_limit:
        await update.message.reply_text(tr(user_id, "wd_min_limit_err", min_limit=min_limit), reply_markup=get_cancel_keyboard(user_id))
        return STATE_WD_AMOUNT

    if amount > max_limit:
        await update.message.reply_text(tr(user_id, "wd_max_limit_err", max_limit=max_limit), reply_markup=get_cancel_keyboard(user_id))
        return STATE_WD_AMOUNT

    if amount > user_data["balance"]:
        await update.message.reply_text(tr(user_id, "wd_insufficient_bal"), reply_markup=get_cancel_keyboard(user_id))
        return STATE_WD_AMOUNT

    number = context.user_data.get("wd_number")
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    fee = 0.0
    if method == "usdt":
        fee = float(get_setting_val("usdt_fee", "0.05"))
        if (amount - fee) <= 0:
            await update.message.reply_text(tr(user_id, "wd_usdt_fee_err"), reply_markup=get_cancel_keyboard(user_id))
            return STATE_WD_AMOUNT

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET balance = balance - ? WHERE user_id = ?", (amount, user_id))
        cursor.execute(
            "INSERT INTO withdrawals (user_id, method, number, amount, status, created_at) VALUES (?, ?, ?, ?, 'Pending', ?)",
            (user_id, method, number, amount, now)
        )
        wd_id = cursor.lastrowid
        conn.commit()

    fee_text = f" (Fee: ৳{fee:.2f})" if method == "usdt" else ""

    await update.message.reply_text(
        tr(user_id, "wd_success_msg", wd_id=wd_id, amount=amount, fee_text=fee_text, method=method.upper(), number=number),
        reply_markup=get_main_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )

    admin_id_setting = get_setting_val("admin_id", "8001997389")
    admins = [int(x.strip()) for x in admin_id_setting.split(",") if x.strip().isdigit()]

    admin_msg = (
        f"💸 **নতুন উইথড্র রিকোয়েস্ট! / New Withdraw Request!**\n\n"
        f"🆔 **উইথড্র আইডি:** #{wd_id}\n"
        f"👤 **ইউজার আইডি:** <code>{user_id}</code>\n"
        f"💵 **পরিমাণ:** ৳{amount:.2f}\n"
        f"📱 **গেটওয়ে:** {method.upper()} - <code>{number}</code>"
    )
    buttons = [
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"adm_app_wd_{wd_id}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"adm_rej_wd_{wd_id}")
        ]
    ]

    for admin in admins:
        try:
            await context.bot.send_message(
                chat_id=admin,
                text=admin_msg,
                reply_markup=InlineKeyboardMarkup(buttons),
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Failed to alert admin of withdraw: {e}")

    return ConversationHandler.END

# -----------------------------------------------------------------------------
# REFERRAL, LEADERBOARD, SUPPORT, LANGUAGE
# -----------------------------------------------------------------------------

async def handle_referrals(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    user_data = get_user_data(user_id)
    if not user_data:
        register_user(update.effective_user)
        user_data = get_user_data(user_id)

    bot_username = (await context.bot.get_me()).username
    ref_link = f"https://t.me/{bot_username}?start={user_id}"
    ref_percent = get_setting_val("ref_bonus_percent", "10.0")

    ref_text = tr(
        user_id, "ref_text",
        ref_link=ref_link,
        ref_count=user_data['ref_count'],
        ref_earnings=user_data['ref_earnings'],
        ref_percent=ref_percent
    )
    await update.message.reply_text(ref_text, reply_markup=get_main_keyboard(user_id), parse_mode=ParseMode.HTML)

async def handle_leaderboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT full_name, balance FROM users WHERE is_banned = 0 ORDER BY balance DESC LIMIT 10")
        top_earners = cursor.fetchall()

    leaderboard = tr(user_id, "leaderboard_title")
    if not top_earners:
        leaderboard += tr(user_id, "no_leaderboard_data")
    else:
        for idx, row in enumerate(top_earners, 1):
            name = html.escape(row["full_name"])
            badge = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else f"{idx}."
            leaderboard += f"{badge} **{name}** — ৳{row['balance']:.2f}\n"

    await update.message.reply_text(leaderboard, reply_markup=get_main_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)

async def handle_support(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    support_handle = get_setting_val("support_handle", "https://t.me/TrustVaultMails_Owners")
    official_channel = get_setting_val("official_channel_link", "https://t.me/TrustVaultMailsOfficial")

    support_text = tr(user_id, "support_text")

    buttons = [
        [InlineKeyboardButton(tr(user_id, "btn_admin_support"), url=support_handle)],
        [InlineKeyboardButton(tr(user_id, "btn_official_channel"), url=official_channel)]
    ]

    await update.message.reply_text(
        support_text,
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.MARKDOWN
    )

async def handle_language(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    is_joined, _ = await check_force_join(user_id, context)
    if not is_joined:
        await start_command(update, context)
        return ConversationHandler.END

    buttons = [
        [InlineKeyboardButton("🇧🇩 বাংলা", callback_data="lang_bn"),
         InlineKeyboardButton("🇺🇸 English", callback_data="lang_en")]
    ]
    await update.message.reply_text(
        tr(user_id, "lang_select_prompt"),
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.MARKDOWN
    )

async def language_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id

    lang = "bn" if data == "lang_bn" else "en"
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET language = ? WHERE user_id = ?", (lang, user_id))
        conn.commit()

    msg = tr(lang, "lang_set_bn") if lang == "bn" else tr(lang, "lang_set_en")
    await query.message.reply_text(msg, reply_markup=get_main_keyboard(user_id, lang=lang), parse_mode=ParseMode.MARKDOWN)

# -----------------------------------------------------------------------------
# ADMIN PANEL HANDLERS
# -----------------------------------------------------------------------------

async def handle_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Pending'")
        pending_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Hold'")
        held_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM withdrawals WHERE status = 'Pending'")
        pending_wd = cursor.fetchone()[0]

    dash_msg = tr(
        user_id, "admin_dash_title",
        total_users=total_users,
        pending_tasks=pending_tasks,
        held_tasks=held_tasks,
        pending_wd=pending_wd
    )

    await update.message.reply_text(
        dash_msg,
        reply_markup=get_admin_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )

async def admin_pending_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM task_submissions WHERE status = 'Pending' ORDER BY created_at DESC LIMIT 10")
        tasks = cursor.fetchall()
    if not tasks:
        await update.message.reply_text(tr(user_id, "admin_no_pending_tasks"), reply_markup=get_admin_keyboard(user_id))
        return
    msg = tr(user_id, "admin_pending_tasks_title")
    keyboard = []
    for task in tasks:
        msg += (
            f"🆔 #{task['submission_id']} | {task['task_type'].upper()}\n"
            f"👤 {html.escape(task['submitted_username'])} | ৳{task['reward_amount']:.2f}\n"
            f"📅 {task['created_at']}\n\n"
        )
        keyboard.append([
            InlineKeyboardButton(f"✅ App #{task['submission_id']}", callback_data=f"adm_app_sub_{task['submission_id']}"),
            InlineKeyboardButton(f"⏳ Hold", callback_data=f"adm_hold_sub_{task['submission_id']}"),
            InlineKeyboardButton(f"❌ Rej", callback_data=f"adm_rej_prompt_{task['submission_id']}")
        ])
    keyboard.append([InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")])
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

async def admin_held_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM task_submissions WHERE status = 'Hold' ORDER BY updated_at DESC LIMIT 15")
        tasks = cursor.fetchall()
    if not tasks:
        await update.message.reply_text(tr(user_id, "admin_no_held_tasks"), reply_markup=get_admin_keyboard(user_id))
        return
    msg = tr(user_id, "admin_held_tasks_title")
    keyboard = []
    for task in tasks:
        msg += (
            f"🆔 <b>Task #{task['submission_id']}</b> | 📱 {task['task_type'].upper()}\n"
            f"👤 <b>User ID:</b> <code>{task['user_id']}</code>\n"
            f"👤 <b>Submission:</b> <code>{html.escape(task['submitted_username'])}</code>\n"
        )
        if task['submitted_password'] and task['submitted_password'] != 'N/A':
            msg += f"🔑 <b>Pass:</b> <code>{html.escape(task['submitted_password'])}</code>\n"
        if task['secret_key_2fa'] and task['secret_key_2fa'] != 'N/A':
            msg += f"🔐 <b>2FA Key:</b> <code>{html.escape(task['secret_key_2fa'])}</code>\n"
        if task['cookies_data']:
            msg += f"🍪 <b>Data:</b> <code>{html.escape(task['cookies_data'][:40])}...</code>\n"
        msg += (
            f"💵 <b>Price:</b> ৳{task['reward_amount']:.2f}\n"
            f"📅 <b>Held At:</b> {task['updated_at']}\n━━━━━━━━━━━━━━━━━━\n"
        )
        keyboard.append([
            InlineKeyboardButton(f"✅ App #{task['submission_id']}", callback_data=f"adm_app_sub_{task['submission_id']}"),
            InlineKeyboardButton(f"❌ Rej #{task['submission_id']}", callback_data=f"adm_rej_prompt_{task['submission_id']}")
        ])
    keyboard.append([InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")])
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

# --- BULK APPROVE & BULK REJECT ---
async def admin_bulk_approve_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    await update.message.reply_text(
        tr(user_id, "admin_bulk_app_prompt"),
        reply_markup=get_cancel_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_ADMIN_BULK_APPROVE

async def admin_bulk_reject_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    await update.message.reply_text(
        tr(user_id, "admin_bulk_rej_prompt"),
        reply_markup=get_cancel_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_ADMIN_BULK_REJECT

async def process_bulk_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    raw_text = update.message.text.strip()
    if is_menu_or_admin_button(raw_text):
        clear_user_state(context)
        return ConversationHandler.END

    items = [x.strip() for x in re.split(r'[\n,\s]+', raw_text) if x.strip()]
    cleaned_items = list(dict.fromkeys(items))

    if not cleaned_items:
        await update.message.reply_text("❌ কোনো ভ্যালিড ডাটা পাওয়া যায়নি।", reply_markup=get_admin_keyboard(user_id))
        return ConversationHandler.END

    approved_count = 0
    matched_count = 0
    unmatched_list = []
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for item in cleaned_items:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM task_submissions WHERE LOWER(submitted_username) = LOWER(?) AND status IN ('Pending', 'Hold')",
                (item,)
            )
            subs = cursor.fetchall()
            
            if not subs:
                unmatched_list.append(item)
                continue

            matched_count += len(subs)

            for sub in subs:
                sub_id = sub["submission_id"]
                sub_user_id = sub["user_id"]
                reward = sub["reward_amount"]

                cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (reward, sub_user_id))
                cursor.execute("UPDATE task_submissions SET status = 'Approved', updated_at = ? WHERE submission_id = ?", (now, sub_id))

                # Lifetime Referral Bonus logic
                cursor.execute("SELECT referred_by FROM users WHERE user_id = ?", (sub_user_id,))
                user_ref = cursor.fetchone()
                if user_ref and user_ref["referred_by"]:
                    referrer_id = user_ref["referred_by"]
                    commission_percent = float(get_setting_val("ref_bonus_percent", "10.0"))
                    commission_bonus = float(reward) * (commission_percent / 100.0)
                    if commission_bonus > 0:
                        cursor.execute(
                            "UPDATE users SET balance = balance + ?, ref_earnings = ref_earnings + ? WHERE user_id = ?",
                            (commission_bonus, commission_bonus, referrer_id)
                        )
                        try:
                            await context.bot.send_message(
                                chat_id=referrer_id,
                                text=tr(referrer_id, "user_ref_commission", bonus=commission_bonus),
                                parse_mode=ParseMode.HTML
                            )
                        except Exception as e:
                            logger.error(f"Failed to notify referrer {referrer_id}: {e}")

                conn.commit()
                approved_count += 1

                sub_username = html.escape(sub['submitted_username'])
                try:
                    await context.bot.send_message(
                        chat_id=sub_user_id,
                        text=tr(sub_user_id, "user_task_approved", task_display=sub["task_type"].upper(), username=sub_username, reward=reward),
                        parse_mode=ParseMode.HTML
                    )
                except Exception as e:
                    logger.error(f"Failed to notify user {sub_user_id}: {e}")

    report_msg = (
        f"✅ **বাল্ক এপ্রুভ সম্পন্ন রিপোর্ট:**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"• ইনপুট দেওয়া মোট আইটেম: `{len(cleaned_items)}` টি\n"
        f"• ডাটাবেজে ম্যাচ পাওয়া টাস্ক: `{matched_count}` টি\n"
        f"• সফলভাবে এপ্রুভ হয়েছে: `{approved_count}` টি\n"
    )
    if unmatched_list:
        report_msg += f"• **পেন্ডিং/হোল্ডে খুঁজে পাওয়া যায়নি ({len(unmatched_list)} টি):**\n`" + ", ".join(unmatched_list[:20]) + ("..." if len(unmatched_list)>20 else "") + "`"

    await update.message.reply_text(
        report_msg,
        reply_markup=get_admin_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    clear_user_state(context)
    return ConversationHandler.END

async def process_bulk_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    raw_text = update.message.text.strip()
    if is_menu_or_admin_button(raw_text):
        clear_user_state(context)
        return ConversationHandler.END

    items = [x.strip() for x in re.split(r'[\n,\s]+', raw_text) if x.strip()]
    cleaned_items = list(dict.fromkeys(items))

    if not cleaned_items:
        await update.message.reply_text("❌ কোনো ভ্যালিড ডাটা পাওয়া যায়নি।", reply_markup=get_admin_keyboard(user_id))
        return ConversationHandler.END

    rejected_count = 0
    matched_count = 0
    unmatched_list = []
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    reason = "Bulk rejection process by admin"

    for item in cleaned_items:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM task_submissions WHERE LOWER(submitted_username) = LOWER(?) AND status IN ('Pending', 'Hold')",
                (item,)
            )
            subs = cursor.fetchall()

            if not subs:
                unmatched_list.append(item)
                continue

            matched_count += len(subs)

            for sub in subs:
                sub_id = sub["submission_id"]
                sub_user_id = sub["user_id"]

                cursor.execute(
                    "UPDATE task_submissions SET status = 'Rejected', reject_reason = ?, updated_at = ? WHERE submission_id = ?",
                    (reason, now, sub_id)
                )
                conn.commit()
                rejected_count += 1

                sub_username = html.escape(sub['submitted_username'])
                esc_reason = html.escape(reason)
                try:
                    await context.bot.send_message(
                        chat_id=sub_user_id,
                        text=tr(sub_user_id, "user_task_rejected", task_display=sub["task_type"].upper(), username=sub_username, reason=esc_reason),
                        parse_mode=ParseMode.HTML
                    )
                except Exception as e:
                    logger.error(f"Failed to notify user {sub_user_id}: {e}")

    report_msg = (
        f"❌ **বাল্ক রিজেক্ট সম্পন্ন রিপোর্ট:**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"• ইনপুট দেওয়া মোট আইটেম: `{len(cleaned_items)}` টি\n"
        f"• ডাটাবেজে ম্যাচ পাওয়া টাস্ক: `{matched_count}` টি\n"
        f"• সফলভাবে রিজেক্ট হয়েছে: `{rejected_count}` টি\n"
    )
    if unmatched_list:
        report_msg += f"• **পেন্ডিং/হোল্ডে খুঁজে পাওয়া যায়নি ({len(unmatched_list)} টি):**\n`" + ", ".join(unmatched_list[:20]) + ("..." if len(unmatched_list)>20 else "") + "`"

    await update.message.reply_text(
        report_msg,
        reply_markup=get_admin_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    clear_user_state(context)
    return ConversationHandler.END

# --- SPREADSHEET UI EXPORT ---
async def admin_excel_export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return

    buttons = [
        [InlineKeyboardButton("📸 IG 2FA (📥 Download)", callback_data="adm_xlsx_ig_2fa"),
         InlineKeyboardButton("🗑️ Clear IG 2FA", callback_data="adm_cls_ig_2fa")],
        [InlineKeyboardButton("📸 IG Seed (📥 Download)", callback_data="adm_xlsx_ig_seed"),
         InlineKeyboardButton("🗑️ Clear IG Seed", callback_data="adm_cls_ig_seed")],
        [InlineKeyboardButton("📘 FB 2FA (📥 Download)", callback_data="adm_xlsx_fb_2fa"),
         InlineKeyboardButton("🗑️ Clear FB 2FA", callback_data="adm_cls_fb_2fa")],
        [InlineKeyboardButton("📘 FB Cookies (📥 Download)", callback_data="adm_xlsx_fb_cookies"),
         InlineKeyboardButton("🗑️ Clear FB Cookies", callback_data="adm_cls_fb_cookies")],
        [InlineKeyboardButton("✉️ Gmail (📥 Download)", callback_data="adm_xlsx_gmail"),
         InlineKeyboardButton("🗑️ Clear Gmail", callback_data="adm_cls_gmail")],
        [InlineKeyboardButton("📊 All Tasks (📥 Download)", callback_data="adm_xlsx_all"),
         InlineKeyboardButton("🗑️ Clear ALL Data", callback_data="adm_cls_all")],
        [InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]
    ]

    await update.message.reply_text(
        "📊 **এক্সেল শীট (.xlsx) ডাউনলোড এবং ক্লিয়ার কন্ট্রোল প্যানেল:**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        "• **Download:** রিজেক্ট হওয়া কাজ বাদ দিয়ে শুধুমাত্র ভ্যালিড কাজের ফাইল ডাউনলোড করবে।\n"
        "• **Clear:** ডাটা আর্কাইভ করবে (রিসেট করবে) যাতে নতুন দিনের কাজ ফ্রেশভাবে ১ থেকে শুরু হয়।",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.MARKDOWN
    )

async def admin_excel_export_download_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        return

    data = query.data.replace("adm_xlsx_", "")

    with get_db() as conn:
        cursor = conn.cursor()
        if data == "all":
            cursor.execute("SELECT * FROM task_submissions WHERE status != 'Rejected' AND is_archived = 0 ORDER BY submission_id DESC")
            task_label = "সকল টাস্ক (All Tasks)"
        else:
            cursor.execute("SELECT * FROM task_submissions WHERE task_type = ? AND status != 'Rejected' AND is_archived = 0 ORDER BY submission_id DESC", (data,))
            task_label = {
                "ig_2fa": "Instagram 2FA",
                "ig_seed": "Instagram Seed",
                "fb_2fa": "Facebook 2FA",
                "fb_cookies": "Facebook Cookies",
                "gmail": "Gmail"
            }.get(data, data.upper())

        tasks = cursor.fetchall()

        if data == "all":
            cursor.execute("SELECT COUNT(*), status FROM task_submissions WHERE status != 'Rejected' AND is_archived = 0 GROUP BY status")
        else:
            cursor.execute("SELECT COUNT(*), status FROM task_submissions WHERE task_type = ? AND status != 'Rejected' AND is_archived = 0 GROUP BY status", (data,))
        status_counts = dict(cursor.fetchall())

        if data == "all":
            cursor.execute("SELECT SUM(reward_amount) FROM task_submissions WHERE status != 'Rejected' AND is_archived = 0")
        else:
            cursor.execute("SELECT SUM(reward_amount) FROM task_submissions WHERE task_type = ? AND status != 'Rejected' AND is_archived = 0", (data,))
        total_amount = cursor.fetchone()[0] or 0.0

    pending = status_counts.get("Pending", 0)
    approved = status_counts.get("Approved", 0)
    hold = status_counts.get("Hold", 0)

    clear_btn_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🗑️ {task_label} ডাটা ক্লিয়ার করুন (0)", callback_data=f"adm_cls_{data}")],
        [InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]
    ])

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = task_label[:30]
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells('A1:J1')
        summary_cell = ws['A1']
        summary_cell.value = f"📊 {task_label.upper()} DASHBOARD | Pending: {pending} | Approved: {approved} | Hold: {hold} | Total Amount: ৳{total_amount:.2f}"
        summary_cell.font = Font(bold=True, size=11, color="FFFFFF")
        summary_cell.fill = PatternFill("solid", fgColor="1F4E78")
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        headers = ["Task ID", "User ID", "Username", "Task Type", "Password", "2FA Secret Key", "Cookies / Data", "Price (BDT)", "Status", "Submitted Time"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 22

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        header_align = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.freeze_panes = 'A4'

        fill_approved = PatternFill("solid", fgColor="E2EFDA")
        fill_pending = PatternFill("solid", fgColor="FFF2CC")
        fill_hold = PatternFill("solid", fgColor="FCE4D6")

        font_approved = Font(color="375623", bold=True)
        font_pending = Font(color="7F6000", bold=True)
        font_hold = Font(color="C65911", bold=True)

        for task in tasks:
            status = task["status"]

            row_data = [
                task["submission_id"],
                task["user_id"],
                task["submitted_username"],
                task["task_type"].upper(),
                task["submitted_password"] or "N/A",
                task["secret_key_2fa"] or "N/A",
                task["cookies_data"] or "N/A",
                task["reward_amount"],
                status,
                task["created_at"]
            ]
            ws.append(row_data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 19

            for col_idx in range(1, 11):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            status_cell = ws.cell(row=current_row, column=9)
            if status == "Approved":
                status_cell.fill = fill_approved
                status_cell.font = font_approved
            elif status == "Pending":
                status_cell.fill = fill_pending
                status_cell.font = font_pending
            elif status == "Hold":
                status_cell.fill = fill_hold
                status_cell.font = font_hold

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        bio.name = f"{data}_valid_tasks_{datetime.date.today()}.xlsx"

        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=bio,
            caption=f"📊 **{task_label} Excel Report (Valid Tasks Only)**\n━━━━━━━━━━━━━━━━━━━━━━\n• **Total Valid Tasks:** {len(tasks)}\n• **Pending:** {pending}\n• **Approved:** {approved}\n• **Hold:** {hold}\n• **Total Amount:** ৳{total_amount:.2f}\n\n*(রিজেক্ট হওয়া কাজগুলো রিপোর্ট থেকে বাদ দেওয়া হয়েছে)*",
            reply_markup=clear_btn_kb,
            parse_mode=ParseMode.MARKDOWN
        )
    else:
        buffer = io.StringIO()
        buffer.write("Task ID,User ID,Username,Task Type,Password,2FA Secret,Cookies,Price,Status,Submitted Time\n")
        for t in tasks:
            buffer.write(f"{t['submission_id']},{t['user_id']},{t['submitted_username']},{t['task_type']},{t['submitted_password']},{t['secret_key_2fa']},{t['cookies_data']},{t['reward_amount']},{t['status']},{t['created_at']}\n")
        
        file_bytes = buffer.getvalue().encode('utf-8')
        bio = io.BytesIO(file_bytes)
        bio.name = f"{data}_tasks.csv"
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=bio,
            caption=f"📊 **{task_label} CSV Report**",
            reply_markup=clear_btn_kb
        )

# --- SOFT DELETE / ARCHIVING CLEAR CALLBACK ---
async def admin_excel_clear_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        return

    data = query.data

    if data.startswith("adm_cls_conf_"):
        cat = data.replace("adm_cls_conf_", "")
        with get_db() as conn:
            cursor = conn.cursor()
            if cat == "all":
                cursor.execute("UPDATE task_submissions SET is_archived = 1 WHERE is_archived = 0")
                label = "সকল কাজের"
            else:
                cursor.execute("UPDATE task_submissions SET is_archived = 1 WHERE task_type = ? AND is_archived = 0", (cat,))
                label = {
                    "ig_2fa": "Instagram 2FA",
                    "ig_seed": "Instagram Seed",
                    "fb_2fa": "Facebook 2FA",
                    "fb_cookies": "Facebook Cookies",
                    "gmail": "Gmail"
                }.get(cat, cat.upper())
            conn.commit()

        await query.edit_message_text(
            f"✅ **{label} সমস্ত ডাটা সফলভাবে ক্লিয়ার (আর্কাইভ) করা হয়েছে!**\n\n"
            f"পূর্বে ডাউনলোডকৃত ডাটা নিরাপদ রয়েছে এবং নতুন জমার কাজ ১ থেকে গণনাকৃত হবে।",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]]),
            parse_mode=ParseMode.MARKDOWN
        )
        return

    if data.startswith("adm_cls_"):
        cat = data.replace("adm_cls_", "")
        label = {
            "ig_2fa": "Instagram 2FA",
            "ig_seed": "Instagram Seed",
            "fb_2fa": "Facebook 2FA",
            "fb_cookies": "Facebook Cookies",
            "gmail": "Gmail",
            "all": "সকল কাজের (ALL)"
        }.get(cat, cat.upper())

        buttons = [
            [InlineKeyboardButton("✅ হ্যাঁ, ক্লিয়ার করুন (Clear 0)", callback_data=f"adm_cls_conf_{cat}")],
            [InlineKeyboardButton("❌ বাতিল", callback_data="adm_back_to_admin")]
        ]

        await query.edit_message_text(
            f"⚠️ **আপনি কি নিশ্চিত যে {label} এর জমার ডাটা ক্লিয়ার (আর্কাইভ) করতে চান?**\n\n"
            f"ডাটা ক্লিয়ার করলে ডাটাবেজে রেকর্ড সেইফ রেখে এক্সেল ফাইল রিসেট হয়ে যাবে।",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.MARKDOWN
        )

# --- USER MANAGEMENT & ADMIN CONTROL PANELS ---
async def admin_user_management(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return

    buttons = [
        [InlineKeyboardButton("🟢 Ban User", callback_data="adm_usr_ban"), InlineKeyboardButton("🟢 Unban User", callback_data="adm_usr_unban")],
        [InlineKeyboardButton("🟢 Add Balance", callback_data="adm_usr_add_bal"), InlineKeyboardButton("🟢 Deduct Balance", callback_data="adm_usr_sub_bal")],
        [InlineKeyboardButton("🟢 Banned List", callback_data="adm_usr_banned_list"), InlineKeyboardButton("🟢 Search User Profile", callback_data="adm_usr_search")],
        [InlineKeyboardButton("🟢 Admin Menu", callback_data="adm_back_to_admin")]
    ]

    await update.message.reply_text(
        "👥 **ইউজার ম্যানেজমেন্ট প্যানেল (User Management):**\n━━━━━━━━━━━━━━━━━━━━━━\nনিচের অপশন থেকে নির্বাচন করুন:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.MARKDOWN
    )

async def admin_control_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return

    force_join = get_setting_val("force_join_enabled", "OFF")

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM sub_admins")
        sub_admins = cursor.fetchall()

    sub_list_str = "\n".join([f"• <code>{row['user_id']}</code>" for row in sub_admins]) if sub_admins else "কোনো সাব-এডমিন নেই।"

    buttons = [
        [InlineKeyboardButton(f"🔒 Force Join: {force_join}", callback_data="adm_toggle_force_join_enabled"),
         InlineKeyboardButton("📢 Edit Force Channels List", callback_data="adm_edit_force_channels_list")],
        [InlineKeyboardButton("➕ Add Sub-Admin", callback_data="adm_add_subadmin"), InlineKeyboardButton("➖ Remove Sub-Admin", callback_data="adm_rem_subadmin")],
        [InlineKeyboardButton("🟢 Admin Menu", callback_data="adm_back_to_admin")]
    ]

    await update.message.reply_text(
        f"👑 **এডমিন কন্ট্রোল প্যানেল (Admin Control):**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👑 **Main Owner ID:** <code>{get_setting_val('admin_id')}</code>\n\n"
        f"👥 **Sub-Admins:**\n{sub_list_str}",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.HTML
    )

async def admin_search_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    context.user_data["admin_search"] = True
    await update.message.reply_text(tr(user_id, "admin_search_prompt_msg"), reply_markup=get_cancel_keyboard(user_id))
    return STATE_ADMIN_SEARCH_QUERY

async def admin_export_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    context.user_data["admin_export"] = True
    await update.message.reply_text(tr(user_id, "admin_export_prompt_msg"), reply_markup=get_cancel_keyboard(user_id))
    return STATE_ADMIN_EXPORT_USER

async def admin_password_manager(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    ig_pass = get_setting_val("ig_default_password", "RBKpass@06")
    fb_pass = get_setting_val("fb_default_password", "FBKpass@07")
    gmail_pass = get_setting_val("gmail_default_password", "aass1122")
    msg = tr(user_id, "admin_pwd_mgr_title", ig_pass=ig_pass, fb_pass=fb_pass, gmail_pass=gmail_pass)
    keyboard = [
        [InlineKeyboardButton("📸 IG Pass", callback_data="adm_edit_ig_default_password")],
        [InlineKeyboardButton("📘 FB Pass", callback_data="adm_edit_fb_default_password")],
        [InlineKeyboardButton("✉️ Gmail Pass", callback_data="adm_edit_gmail_default_password")],
        [InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]
    ]
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

async def admin_price_manager(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    ig_2fa = get_setting_val("ig_2fa_price", "4.00")
    ig_seed = get_setting_val("ig_seed_price", "4.30")
    fb_2fa = get_setting_val("fb_2fa_price", "5.00")
    fb_cookies = get_setting_val("fb_cookies_price", "7.00")
    gmail = get_setting_val("gmail_task_price", "22.00")
    ref_bonus = get_setting_val("ref_bonus_percent", "10.0")

    msg = tr(
        user_id, "admin_price_mgr_title",
        ig_2fa=ig_2fa, ig_seed=ig_seed, fb_2fa=fb_2fa, fb_cookies=fb_cookies, gmail=gmail, ref_bonus=ref_bonus
    )
    keyboard = [
        [InlineKeyboardButton(f"📸 IG 2FA (৳{ig_2fa})", callback_data="adm_edit_ig_2fa_price"),
         InlineKeyboardButton(f"📸 IG Seed (৳{ig_seed})", callback_data="adm_edit_ig_seed_price")],
        [InlineKeyboardButton(f"📘 FB 2FA (৳{fb_2fa})", callback_data="adm_edit_fb_2fa_price"),
         InlineKeyboardButton(f"📘 FB Cookies (৳{fb_cookies})", callback_data="adm_edit_fb_cookies_price")],
        [InlineKeyboardButton(f"✉️ Gmail (৳{gmail})", callback_data="adm_edit_gmail_task_price"),
         InlineKeyboardButton(f"👥 Ref Commission ({ref_bonus}%)", callback_data="adm_edit_ref_bonus_percent")],
        [InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]
    ]
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

async def admin_withdraw_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM withdrawals WHERE status = 'Pending' ORDER BY created_at DESC LIMIT 10")
        wds = cursor.fetchall()
    if not wds:
        await update.message.reply_text(tr(user_id, "admin_no_pending_wds"), reply_markup=get_admin_keyboard(user_id))
        return
    msg = tr(user_id, "admin_pending_wds_title")
    keyboard = []
    for wd in wds:
        msg += f"🆔 #{wd['id']} | {wd['method'].upper()} | ৳{wd['amount']:.2f} | <code>{wd['number']}</code>\n"
        keyboard.append([
            InlineKeyboardButton(f"✅ App #{wd['id']}", callback_data=f"adm_app_wd_{wd['id']}"),
            InlineKeyboardButton(f"❌ Rej #{wd['id']}", callback_data=f"adm_rej_wd_{wd['id']}")
        ])
    keyboard.append([InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")])
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

async def admin_system_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    force_join = get_setting_val("force_join_enabled", "OFF")
    maint = get_setting_val("maintenance_mode", "OFF")
    ig_2fa = get_setting_val("ig_2fa_active", "ON")
    ig_seed = get_setting_val("ig_seed_active", "ON")
    fb_2fa = get_setting_val("fb_2fa_active", "ON")
    fb_cookies = get_setting_val("fb_cookies_active", "ON")
    gmail = get_setting_val("gmail_task_active", "ON")
    
    bkash_wd = get_setting_val("wd_bkash_active", "ON")
    nagad_wd = get_setting_val("wd_nagad_active", "ON")
    usdt_wd = get_setting_val("wd_usdt_active", "ON")

    min_bkash = get_setting_val("min_withdraw_bkash", "50.0")
    max_bkash = get_setting_val("max_withdraw_bkash", "5000.0")
    min_nagad = get_setting_val("min_withdraw_nagad", "50.0")
    max_nagad = get_setting_val("max_withdraw_nagad", "5000.0")
    min_usdt = get_setting_val("min_withdraw_usdt", "0.25")
    max_usdt = get_setting_val("max_withdraw_usdt", "100.0")

    msg = tr(
        user_id, "admin_sys_settings_title",
        force_join=force_join, maint=maint, ig_2fa=ig_2fa, ig_seed=ig_seed,
        fb_2fa=fb_2fa, fb_cookies=fb_cookies, gmail=gmail,
        bkash_wd=bkash_wd, nagad_wd=nagad_wd, usdt_wd=usdt_wd,
        min_bkash=min_bkash, max_bkash=max_bkash,
        min_nagad=min_nagad, max_nagad=max_nagad,
        min_usdt=min_usdt, max_usdt=max_usdt
    )
    keyboard = [
        [InlineKeyboardButton(f"Force Join: {force_join}", callback_data="adm_toggle_force_join_enabled"),
         InlineKeyboardButton("📢 Edit Force Channels", callback_data="adm_edit_force_channels_list")],
        [InlineKeyboardButton(f"Maintenance: {maint}", callback_data="adm_toggle_maintenance_mode")],
        [InlineKeyboardButton(f"🟢 Bkash: {bkash_wd}", callback_data="adm_toggle_wd_bkash_active"),
         InlineKeyboardButton("✏️ Bkash Min/Max", callback_data="adm_edit_bkash_limits")],
        [InlineKeyboardButton(f"🟠 Nagad: {nagad_wd}", callback_data="adm_toggle_wd_nagad_active"),
         InlineKeyboardButton("✏️ Nagad Min/Max", callback_data="adm_edit_nagad_limits")],
        [InlineKeyboardButton(f"🔵 USDT: {usdt_wd}", callback_data="adm_toggle_wd_usdt_active"),
         InlineKeyboardButton("✏️ USDT Min/Max", callback_data="adm_edit_usdt_limits")],
        [InlineKeyboardButton(f"📸 IG 2FA: {ig_2fa}", callback_data="adm_toggle_ig_2fa_active"),
         InlineKeyboardButton(f"📸 IG Seed: {ig_seed}", callback_data="adm_toggle_ig_seed_active")],
        [InlineKeyboardButton(f"📘 FB 2FA: {fb_2fa}", callback_data="adm_toggle_fb_2fa_active"),
         InlineKeyboardButton(f"📘 FB Cookies: {fb_cookies}", callback_data="adm_toggle_fb_cookies_active")],
        [InlineKeyboardButton(f"✉️ Gmail: {gmail}", callback_data="adm_toggle_gmail_task_active")],
        [InlineKeyboardButton(tr(user_id, "btn_admin_menu"), callback_data="adm_back_to_admin")]
    ]
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.HTML)

async def admin_live_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Pending'")
        pending_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Hold'")
        held_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Approved'")
        approved_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM task_submissions WHERE status = 'Rejected'")
        rejected_tasks = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM withdrawals WHERE status = 'Pending'")
        pending_wd = cursor.fetchone()[0]
        cursor.execute("SELECT SUM(balance) FROM users")
        total_balance = cursor.fetchone()[0] or 0.0

    msg = tr(
        user_id, "admin_live_stats_title",
        total_users=total_users,
        pending_tasks=pending_tasks,
        held_tasks=held_tasks,
        approved_tasks=approved_tasks,
        rejected_tasks=rejected_tasks,
        pending_wd=pending_wd,
        total_balance=total_balance
    )
    await update.message.reply_text(msg, reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)

async def admin_broadcast_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    context.user_data["admin_broadcast"] = True
    await update.message.reply_text(
        tr(user_id, "admin_broadcast_prompt_msg"),
        reply_markup=get_cancel_keyboard(user_id),
        parse_mode=ParseMode.MARKDOWN
    )
    return STATE_ADMIN_BROADCAST

# -----------------------------------------------------------------------------
# ADMIN CALLBACK QUERY HANDLER
# -----------------------------------------------------------------------------
async def admin_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        await query.edit_message_text("❌ You are not authorized.")
        return

    data = query.data

    if data == "adm_back_to_admin":
        await query.edit_message_text(
            "🔑 **Returning to Admin Panel...**",
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    if data.startswith("adm_toggle_"):
        key = data.replace("adm_toggle_", "")
        current = get_setting_val(key, "OFF")
        new_val = "ON" if current == "OFF" else "OFF"
        set_setting_val(key, new_val)
        await query.edit_message_text(f"✅ ````{key}```` toggled to ````{new_val}````.")
        return

    if data == "adm_edit_force_channels_list":
        current_list = get_setting_val(
            "force_channels_list",
            "@TrustVaultMailsOfficial|https://t.me/TrustVaultMailsOfficial"
        )
        await query.edit_message_text(
            "📢 <b>ফোর্স জয়েন চ্যানেল লিস্ট সেট করুন</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"
            "প্রতি লাইনে একটি করে চ্যানেল দিন।\n\n"
            "<b>Public channel:</b> <code>@channelusername|https://t.me/channelusername</code>\n"
            "<b>Private channel:</b> <code>--1004397984847|https://t.me/+invite_hash</code>\n\n"
            "⚠️ Private channel-এর ক্ষেত্রে বাম পাশে অবশ্যই আসল <b>Chat ID</b> দিতে হবে এবং bot-কে ওই channel-এর administrator করতে হবে।\n\n"
            f"<b>বর্তমান লিস্ট:</b>\n<code>{html.escape(current_list)}</code>\n\n"
            "👉 নতুন লিস্ট নিচে পাঠান:",
            parse_mode=ParseMode.HTML
        )
        return STATE_ADMIN_FORCE_CHANNELS_EDIT

    if data in ["adm_edit_bkash_limits", "adm_edit_nagad_limits", "adm_edit_usdt_limits"]:
        m_name = data.replace("adm_edit_", "").replace("_limits", "")
        context.user_data["edit_limits_method"] = m_name
        current_min = get_setting_val(f"min_withdraw_{m_name}", "50.0")
        current_max = get_setting_val(f"max_withdraw_{m_name}", "5000.0")
        await query.edit_message_text(
            f"✏️ **{m_name.upper()} উইথড্র লিমিট সেট করুন:**\n\n"
            f"বর্তমান সর্বনিম্ন: `{current_min}` | সর্বোচ্চ: `{current_max}`\n\n"
            f"পাঠানোর ফরম্যাট: `সর্বনিম্ন সর্বোচ্চ` (যেমন: `50 5000`)",
            parse_mode=ParseMode.MARKDOWN
        )
        return STATE_ADMIN_SETTING_VAL

    if data.startswith("adm_edit_"):
        setting_key = data.replace("adm_edit_", "")
        context.user_data["edit_setting_key"] = setting_key
        await query.edit_message_text(f"✏️ Send new value for **{setting_key}**:")
        return STATE_ADMIN_SETTING_VAL

    # Sub Admin Controls
    if data == "adm_add_subadmin":
        if not is_owner(user_id):
            await query.answer("❌ শুধুমাত্র প্রধান মালিক সাব-এডমিন যোগ করতে পারবেন!", show_alert=True)
            return
        await query.edit_message_text("➕ **নতুন সাব-এডমিনের টেলিগ্রাম ইউজার আইডি পাঠাবে:**")
        return STATE_ADMIN_ADD_SUBADMIN

    if data == "adm_rem_subadmin":
        if not is_owner(user_id):
            await query.answer("❌ শুধুমাত্র প্রধান মালিক সাব-এডমিন রিমুভ করতে পারবেন!", show_alert=True)
            return
        await query.edit_message_text("➖ **রিমুভ করার জন্য সাব-এডমিনের টেলিগ্রাম ইউজার আইডি পাঠাবে:**")
        return STATE_ADMIN_REM_SUBADMIN

    # User Management Callbacks
    if data == "adm_usr_ban":
        await query.edit_message_text("🚫 **ব্যান করার জন্য ইউজারের টেলিগ্রাম আইডি পাঠান:**")
        return STATE_ADMIN_BAN_USER

    if data == "adm_usr_unban":
        await query.edit_message_text("✅ **আনব্যান করার জন্য ইউজারের টেলিগ্রাম আইডি পাঠান:**")
        return STATE_ADMIN_UNBAN_USER

    if data == "adm_usr_add_bal":
        await query.edit_message_text(
            "💰 **ইউজারের ওয়ালেটে ব্যালেন্স যোগ করুন:**\n\nফরম্যাট: `ইউজার_আইডি পরিমাণ` (যেমন: `8001997389 50`)",
            parse_mode=ParseMode.MARKDOWN
        )
        return STATE_ADMIN_ADD_BAL

    if data == "adm_usr_sub_bal":
        await query.edit_message_text(
            "💸 **ইউজারের ওয়ালেট থেকে ব্যালেন্স কাটুন:**\n\nফরম্যাট: `ইউজার_আইডি পরিমাণ` (যেমন: `8001997389 20`)",
            parse_mode=ParseMode.MARKDOWN
        )
        return STATE_ADMIN_SUB_BAL

    if data == "adm_usr_banned_list":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id, username, full_name FROM users WHERE is_banned = 1")
            banned = cursor.fetchall()
        if not banned:
            msg = "📜 **কোনো ব্যান ইউজার পাওয়া যায়নি।**"
        else:
            msg = "📜 **নিষিদ্ধ (Banned) ইউজার লিস্ট:**\n━━━━━━━━━━━━━━━━━━━━━━\n"
            for u in banned:
                uname = f"@{u['username']}" if u['username'] else "No Username"
                msg += f"• 🆔 ID: <code>{u['user_id']}</code> | Username: {uname} | Name: {html.escape(u['full_name'] or 'N/A')}\n"
        
        buttons = [[InlineKeyboardButton("🟢 Admin Menu", callback_data="adm_back_to_admin")]]
        await query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        return

    if data == "adm_usr_search":
        await query.edit_message_text("🔍 **প্রোফাইল সার্চ করতে ইউজারের Telegram ID বা Username লিখুন:**")
        return STATE_ADMIN_USER_SEARCH

    # Approve Task (With Lifetime Referral Commission)
    if data.startswith("adm_app_sub_"):
        sub_id = int(data.replace("adm_app_sub_", ""))
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM task_submissions WHERE submission_id = ?", (sub_id,))
            sub = cursor.fetchone()
            if not sub or sub["status"] == "Approved":
                await query.edit_message_text("❌ Task already approved or processed!")
                return

            cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (sub["reward_amount"], sub["user_id"]))
            cursor.execute("UPDATE task_submissions SET status = 'Approved', updated_at = ? WHERE submission_id = ?", (now, sub_id))

            cursor.execute("SELECT referred_by FROM users WHERE user_id = ?", (sub["user_id"],))
            user_ref = cursor.fetchone()
            if user_ref and user_ref["referred_by"]:
                referrer_id = user_ref["referred_by"]
                commission_percent = float(get_setting_val("ref_bonus_percent", "10.0"))
                commission_bonus = float(sub["reward_amount"]) * (commission_percent / 100.0)
                if commission_bonus > 0:
                    cursor.execute(
                        "UPDATE users SET balance = balance + ?, ref_earnings = ref_earnings + ? WHERE user_id = ?",
                        (commission_bonus, commission_bonus, referrer_id)
                    )
                    try:
                        await context.bot.send_message(
                            chat_id=referrer_id,
                            text=tr(referrer_id, "user_ref_commission", bonus=commission_bonus),
                            parse_mode=ParseMode.HTML
                        )
                    except Exception as e:
                        logger.error(f"Error notifying referrer {referrer_id}: {e}")
            conn.commit()

        task_display = sub["task_type"].upper()
        sub_username = html.escape(sub['submitted_username'])

        if sub["task_type"].lower() == "gmail":
            with get_db() as balance_conn:
                balance_cursor = balance_conn.cursor()
                balance_cursor.execute("SELECT balance FROM users WHERE user_id = ?", (sub["user_id"],))
                balance_row = balance_cursor.fetchone()
            user_balance = balance_row["balance"] if balance_row else 0
            amount = sub["reward_amount"]
            text = (
                "🎉 <b>টাস্ক সফলভাবে এপ্রুভ হয়েছে!</b>\n\n"
                "📋 <b>টাস্ক:</b> 📧 Gmail Sell\n"
                "🔹 <b>স্ট্যাটাস:</b> ✅ Approved\n"
                f"💰 <b>যোগকৃত অর্থ:</b> ৳{amount:.2f}\n"
                f"💳 <b>বর্তমান ব্যালেন্স:</b> ৳{user_balance:.2f}\n\n"
                "<i>TrustVault Mails-এর সাথে থাকার জন্য ধন্যবাদ!</i>"
            )
        else:
            text = tr(sub["user_id"], "user_task_approved", task_display=task_display, username=sub_username, reward=sub['reward_amount'])

        try:
            await context.bot.send_message(
                chat_id=sub["user_id"], text=text, parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Error sending approval DM to user {sub['user_id']}: {e}")

        await query.edit_message_text(f"✅ Task #{sub_id} successfully approved!")
        return

    # Hold Task
    if data.startswith("adm_hold_sub_"):
        sub_id = int(data.replace("adm_hold_sub_", ""))
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE task_submissions SET status = 'Hold', updated_at = ? WHERE submission_id = ?", (now, sub_id))
            cursor.execute("SELECT user_id FROM task_submissions WHERE submission_id = ?", (sub_id,))
            row = cursor.fetchone()
            conn.commit()

        if row:
            target_user_id = row["user_id"]
            try:
                await context.bot.send_message(
                    chat_id=target_user_id,
                    text=tr(target_user_id, "user_task_held", sub_id=sub_id),
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                logger.error(f"Failed to notify user {target_user_id} of hold: {e}")

        await query.edit_message_text(f"⌛ Task #{sub_id} placed on Hold.")
        return

    # Reject Prompt
    if data.startswith("adm_rej_prompt_"):
        sub_id = int(data.replace("adm_rej_prompt_", ""))
        buttons = [
            [InlineKeyboardButton("📋 Default Reason", callback_data=f"adm_rej_act_{sub_id}_default")],
            [InlineKeyboardButton("🔑 Wrong Password", callback_data=f"adm_rej_act_{sub_id}_wrongpass")],
            [InlineKeyboardButton("📝 Custom Reason", callback_data=f"adm_rej_custom_{sub_id}")]
        ]
        await query.edit_message_text("❌ Select rejection reason:", reply_markup=InlineKeyboardMarkup(buttons))
        return

    # Reject Action
    if data.startswith("adm_rej_act_"):
        parts = data.replace("adm_rej_act_", "").split("_")
        sub_id = int(parts[0])
        reason_key = parts[1]
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if reason_key == "default":
            reason = "Account issue detected during review"
        elif reason_key == "wrongpass":
            reason = "Wrong password or username mismatch"
        else:
            reason = "Submission criteria not met"

        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE task_submissions SET status = 'Rejected', reject_reason = ?, updated_at = ? WHERE submission_id = ?", (reason, now, sub_id))
            cursor.execute("SELECT * FROM task_submissions WHERE submission_id = ?", (sub_id,))
            sub = cursor.fetchone()
            conn.commit()

        if sub:
            task_display = sub["task_type"].upper()
            sub_username = html.escape(sub['submitted_username'])
            esc_reason = html.escape(reason)

            if sub["task_type"].lower() == "gmail":
                text = (
                    "❌ <b>টাস্কটি গ্রহণযোগ্য হয়নি!</b>\n\n"
                    "📋 <b>টাস্ক:</b> 📧 Gmail Sell\n"
                    "🔹 <b>স্ট্যাটাস:</b> ❌ Rejected\n"
                    f"⚠️ <b>কারণ:</b> {esc_reason}\n\n"
                    "<i>সঠিক তথ্য দিয়ে পুনরায় চেষ্টা করুন।</i>"
                )
            else:
                text = tr(sub["user_id"], "user_task_rejected", task_display=task_display, username=sub_username, reason=esc_reason)

            try:
                await context.bot.send_message(
                    chat_id=sub["user_id"], text=text, parse_mode=ParseMode.HTML
                )
            except Exception as e:
                logger.error(f"Error sending reject DM to user {sub['user_id']}: {e}")

        await query.edit_message_text(f"❌ Task #{sub_id} rejected.")
        return

    # Approve Withdraw
    if data.startswith("adm_app_wd_"):
        wd_id = int(data.replace("adm_app_wd_", ""))
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM withdrawals WHERE id = ?", (wd_id,))
            wd = cursor.fetchone()
            if not wd or wd["status"] != "Pending":
                await query.edit_message_text("❌ Already approved or processed!")
                return
            cursor.execute("UPDATE withdrawals SET status = 'Approved' WHERE id = ?", (wd_id,))
            conn.commit()
        try:
            await context.bot.send_message(
                chat_id=wd["user_id"],
                text=tr(wd["user_id"], "user_wd_approved", wd_id=wd_id, amount=wd['amount'], number=wd['number']),
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Failed to notify user of withdrawal approval: {e}")

        await query.edit_message_text(f"✅ Withdraw #{wd_id} (৳{wd['amount']}) approved!")
        return

    # Reject Withdraw
    if data.startswith("adm_rej_wd_"):
        wd_id = int(data.replace("adm_rej_wd_", ""))
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM withdrawals WHERE id = ?", (wd_id,))
            wd = cursor.fetchone()
            if not wd or wd["status"] != "Pending":
                await query.edit_message_text("❌ Already rejected or processed!")
                return
            cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (wd["amount"], wd["user_id"]))
            cursor.execute("UPDATE withdrawals SET status = 'Rejected' WHERE id = ?", (wd_id,))
            conn.commit()
        try:
            await context.bot.send_message(
                chat_id=wd["user_id"],
                text=tr(wd["user_id"], "user_wd_rejected", wd_id=wd_id, amount=wd['amount']),
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Failed to notify user of withdrawal rejection: {e}")

        await query.edit_message_text(f"❌ Withdraw #{wd_id} (৳{wd['amount']}) rejected & refunded to user!")
        return

    if data.startswith("adm_rej_custom_"):
        sub_id = int(data.replace("adm_rej_custom_", ""))
        context.user_data["rej_sub_id"] = sub_id
        await query.edit_message_text(tr(user_id, "admin_rej_custom_prompt"))
        return STATE_ADMIN_REJECT_REASON

# -----------------------------------------------------------------------------
# HIGH-SPEED ASYNC BATCH BROADCAST ENGINE
# -----------------------------------------------------------------------------
async def process_admin_broadcast_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    msg = update.message

    if msg.text:
        text_str = msg.text.strip()
        if text_str in ["❌ বাতিল", "❌ Cancel", "🔙 মূল মেনু", "🟢 মূল মেনু", "🔙 Main Menu", "🟢 Main Menu"]:
            clear_user_state(context)
            await update.message.reply_text(
                tr(user_id, "operation_cancelled"),
                reply_markup=get_admin_keyboard(user_id)
            )
            return ConversationHandler.END

    if msg.text:
        msg_type_str = "📝 Text"
    elif msg.photo:
        msg_type_str = "🖼️ Photo"
    elif msg.video:
        msg_type_str = "🎥 Video"
    elif msg.document:
        msg_type_str = "📄 Document"
    elif msg.audio:
        msg_type_str = "🎵 Audio"
    elif msg.voice:
        msg_type_str = "🎤 Voice"
    elif msg.sticker:
        msg_type_str = "🎨 Sticker"
    elif msg.video_note:
        msg_type_str = "📹 Video Note"
    else:
        msg_type_str = "📦 Media"

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM users WHERE is_banned = 0")
        users = cursor.fetchall()

    total_users = len(users)
    start_time = time.time()

    status_msg = await update.message.reply_text(
        f"🚀 **High-Speed Broadcast Started...**\n\n"
        f"👥 **Total Target Users:** {total_users}\n"
        f"📨 **Type:** {msg_type_str}\n"
        f"⏳ **Status:** Preparing batch send...",
        parse_mode=ParseMode.MARKDOWN
    )

    success_count = 0
    blocked_count = 0
    deleted_count = 0
    other_failed_count = 0

    semaphore = asyncio.Semaphore(25)

    async def send_to_user(target_id):
        nonlocal success_count, blocked_count, deleted_count, other_failed_count
        async with semaphore:
            for attempt in range(2):
                try:
                    await context.bot.copy_message(
                        chat_id=target_id,
                        from_chat_id=update.effective_chat.id,
                        message_id=msg.message_id
                    )
                    success_count += 1
                    return
                except RetryAfter as e:
                    await asyncio.sleep(e.retry_after + 0.2)
                except Forbidden:
                    blocked_count += 1
                    return
                except (BadRequest, TelegramError) as e:
                    err_str = str(e).lower()
                    if "chat not found" in err_str or "user is deactivated" in err_str or "bot was blocked" in err_str:
                        deleted_count += 1
                    else:
                        other_failed_count += 1
                    return
                except Exception:
                    other_failed_count += 1
                    return

    chunk_size = 50
    user_ids = [u["user_id"] for u in users]
    
    for i in range(0, total_users, chunk_size):
        chunk = user_ids[i:i + chunk_size]
        await asyncio.gather(*[send_to_user(uid) for uid in chunk])
        
        processed = min(i + chunk_size, total_users)
        if processed % 200 == 0 or processed == total_users:
            try:
                elapsed = round(time.time() - start_time, 1)
                await status_msg.edit_text(
                    f"⚡ **High-Speed Broadcast Processing...**\n\n"
                    f"👥 **Total:** {total_users}\n"
                    f"⏳ **Processed:** {processed}/{total_users}\n"
                    f"✅ **Sent:** {success_count}\n"
                    f"❌ **Failed:** {blocked_count + deleted_count + other_failed_count}\n"
                    f"⏱️ **Time:** {elapsed}s",
                    parse_mode=ParseMode.MARKDOWN
                )
            except Exception:
                pass
        
        await asyncio.sleep(0.05)

    total_failed = blocked_count + deleted_count + other_failed_count
    success_rate = round((success_count / total_users * 100), 1) if total_users > 0 else 0.0
    elapsed_total = round(time.time() - start_time, 1)

    final_report = tr(
        user_id, "admin_broadcast_done",
        total=total_users,
        msg_type=msg_type_str,
        count=success_count,
        failed=total_failed,
        success_rate=success_rate,
        time_taken=elapsed_total,
        blocked=blocked_count,
        deleted=deleted_count,
        other=other_failed_count
    )

    try:
        await status_msg.edit_text(final_report, parse_mode=ParseMode.MARKDOWN)
    except Exception:
        await update.message.reply_text(final_report, parse_mode=ParseMode.MARKDOWN)

    await update.message.reply_text(
        "🔙 **অ্যাডমিন প্যানেল মেনু:**",
        reply_markup=get_admin_keyboard(user_id)
    )

    clear_user_state(context)
    return ConversationHandler.END

# -----------------------------------------------------------------------------
# PROCESS ADMIN TEXT INPUTS
# -----------------------------------------------------------------------------
async def process_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()

    if context.user_data.get("admin_search"):
        with get_db() as conn:
            cursor = conn.cursor()
            if text.isdigit():
                cursor.execute("SELECT * FROM task_submissions WHERE submission_id = ? OR user_id = ?", (int(text), int(text)))
            else:
                cursor.execute("SELECT * FROM task_submissions WHERE submitted_username LIKE ?", (f"%{text}%",))
            results = cursor.fetchall()

        if not results:
            await update.message.reply_text(tr(user_id, "admin_search_no_results"), reply_markup=get_admin_keyboard(user_id))
        else:
            msg = tr(user_id, "admin_search_results_title")
            for r in results[:10]:
                msg += f"🆔 #{r['submission_id']} | {r['task_type'].upper()} | {html.escape(r['submitted_username'])} | Status: {r['status']}\n"
            await update.message.reply_text(msg, reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.HTML)

        context.user_data["admin_search"] = False
        return ConversationHandler.END

    if context.user_data.get("admin_export"):
        if not text.isdigit():
            await update.message.reply_text(tr(user_id, "admin_export_invalid_id"), reply_markup=get_cancel_keyboard(user_id))
            return STATE_ADMIN_EXPORT_USER

        target_id = int(text)
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM task_submissions WHERE user_id = ?", (target_id,))
            subs = cursor.fetchall()

        if not subs:
            await update.message.reply_text(tr(user_id, "admin_export_no_tasks"), reply_markup=get_admin_keyboard(user_id))
        else:
            buffer = io.StringIO()
            buffer.write(f"Trust Vault Mails STYLE - USER {target_id} REPORT\n")
            buffer.write("="*50+"\n")
            for s in subs:
                buffer.write(f"ID:{s['submission_id']} Type:{s['task_type']} User:{s['submitted_username']} Pass:{s['submitted_password']} Status:{s['status']}\n")

            file_bytes = buffer.getvalue().encode('utf-8')
            bio = io.BytesIO(file_bytes)
            bio.name = f"user_{target_id}.txt"
            await context.bot.send_document(chat_id=update.effective_chat.id, document=bio, caption=f"📂 Report User {target_id}")

        context.user_data["admin_export"] = False
        return ConversationHandler.END

    if context.user_data.get("rej_sub_id"):
        sub_id = context.user_data["rej_sub_id"]
        reason = text
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE task_submissions SET status = 'Rejected', reject_reason = ?, updated_at = ? WHERE submission_id = ?", (reason, now, sub_id))
            cursor.execute("SELECT * FROM task_submissions WHERE submission_id = ?", (sub_id,))
            sub = cursor.fetchone()
            conn.commit()

        if sub:
            task_display = sub["task_type"].upper()
            sub_username = html.escape(sub['submitted_username'])
            esc_reason = html.escape(reason)
            try:
                await context.bot.send_message(
                    chat_id=sub["user_id"],
                    text=tr(sub["user_id"], "user_task_rejected", task_display=task_display, username=sub_username, reason=esc_reason),
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                logger.error(f"Error sending custom reject DM to user {sub['user_id']}: {e}")

        await update.message.reply_text(f"❌ Task #{sub_id} rejected.", reply_markup=get_admin_keyboard(user_id))
        context.user_data["rej_sub_id"] = None
        return ConversationHandler.END

    if context.user_data.get("edit_limits_method"):
        m_name = context.user_data["edit_limits_method"]
        parts = text.split()
        if len(parts) != 2:
            await update.message.reply_text("❌ সঠিক ফরম্যাটে পাঠাবে (যেমন: `50 5000`)")
            return STATE_ADMIN_SETTING_VAL

        try:
            mn = float(parts[0])
            mx = float(parts[1])
            set_setting_val(f"min_withdraw_{m_name}", str(mn))
            set_setting_val(f"max_withdraw_{m_name}", str(mx))
            await update.message.reply_text(f"✅ **{m_name.upper()}** উইথড্র লিমিট সেট হয়েছে:\n• Min: ৳{mn}\n• Max: ৳{mx}", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
        except ValueError:
            await update.message.reply_text("❌ সঠিক সংখ্যা লিখুন।")
            return STATE_ADMIN_SETTING_VAL

        context.user_data["edit_limits_method"] = None
        return ConversationHandler.END

    if context.user_data.get("edit_setting_key"):
        key = context.user_data["edit_setting_key"]
        set_setting_val(key, text)
        await update.message.reply_text(f"✅ ````{key}```` updated to ````{text}````.", reply_markup=get_admin_keyboard(user_id))
        context.user_data["edit_setting_key"] = None
        return ConversationHandler.END

    await update.message.reply_text("🔙 **অ্যাডমিন প্যানেল মেনু:**", reply_markup=get_admin_keyboard(user_id))
    return ConversationHandler.END

async def process_admin_force_channels_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    if is_menu_or_admin_button(text):
        clear_user_state(context)
        return ConversationHandler.END

    channels = parse_force_join_channels(text)
    if not channels:
        await update.message.reply_text(
            "❌ কোনো বৈধ channel পাওয়া যায়নি।\n\n"
            "উদাহরণ: <code>@mychannel|https://t.me/mychannel</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=get_cancel_keyboard(user_id)
        )
        return STATE_ADMIN_FORCE_CHANNELS_EDIT

    # Store a clean one-entry-per-line representation.
    normalized = []
    for channel_id, join_link in channels:
        normalized.append(f"{channel_id}|{join_link}")
    set_setting_val("force_channels_list", "\n".join(normalized))
    await update.message.reply_text(
        "✅ <b>ফোর্স জয়েন চ্যানেল লিস্ট সফলভাবে আপডেট হয়েছে!</b>\n\n"
        f"📢 Configured channels: <b>{len(normalized)}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=get_admin_keyboard(user_id)
    )
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_user_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip().replace("@", "")
    if is_menu_or_admin_button(text):
        clear_user_state(context)
        return ConversationHandler.END

    with get_db() as conn:
        cursor = conn.cursor()
        if text.isdigit():
            cursor.execute("SELECT * FROM users WHERE user_id = ? OR LOWER(username) = LOWER(?)", (int(text), text))
        else:
            cursor.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (text,))
        u = cursor.fetchone()

    buttons = [[InlineKeyboardButton("🟢 Admin Menu", callback_data="adm_back_to_admin")]]

    if not u:
        await update.message.reply_text(
            "❌ **ইউজার পাওয়া যায়নি!**\nসঠিক Telegram User ID অথবা Username লিখুন।",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.MARKDOWN
        )
    else:
        ban_str = "🚫 Banned" if u["is_banned"] else "✅ Active"
        uname_str = f"@{u['username']}" if u["username"] else "N/A"
        msg = (
            f"👤 **User Profile Details:**\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🆔 **User ID:** <code>{u['user_id']}</code>\n"
            f"👤 **Username:** {uname_str}\n"
            f"📝 **Full Name:** {html.escape(u['full_name'] or 'N/A')}\n"
            f"💰 **Balance:** ৳{u['balance']:.2f}\n"
            f"🚫 **Ban Status:** {ban_str}\n"
            f"📅 **Joined Date:** {u['joined_at'] or 'N/A'}\n"
            f"👥 **Referral Count:** {u['ref_count']} জন\n"
            f"💵 **Referral Earnings:** ৳{u['ref_earnings']:.2f}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━"
        )
        await update.message.reply_text(
            msg,
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )

    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_add_subadmin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_owner(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ সঠিক টেলিগ্রাম ইউজার আইডি পাঠাবে।")
        return STATE_ADMIN_ADD_SUBADMIN

    sub_id = int(text)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT OR REPLACE INTO sub_admins (user_id, added_by, created_at) VALUES (?, ?, ?)", (sub_id, user_id, now))
        conn.commit()

    await update.message.reply_text(f"✅ ইউজার আইডি ````{sub_id}```` সফলভাবে সাব-এডমিন হিসেবে যুক্ত করা হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_rem_subadmin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_owner(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ সঠিক টেলিগ্রাম ইউজার আইডি পাঠাবে।")
        return STATE_ADMIN_REM_SUBADMIN

    sub_id = int(text)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM sub_admins WHERE user_id = ?", (sub_id,))
        conn.commit()

    await update.message.reply_text(f"✅ সাব-এডমিন ````{sub_id}```` সফলভাবে রিমুভ করা হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_ban_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ সঠিক টেলিগ্রাম ইউজার আইডি পাঠাবে।")
        return STATE_ADMIN_BAN_USER

    target_id = int(text)
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET is_banned = 1 WHERE user_id = ?", (target_id,))
        conn.commit()

    await update.message.reply_text(f"🚫 ইউজার ````{target_id}```` কে সিস্টেমে ব্যান করা হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_unban_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ সঠিক টেলিগ্রাম ইউজার আইডি পাঠাবে।")
        return STATE_ADMIN_UNBAN_USER

    target_id = int(text)
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET is_banned = 0 WHERE user_id = ?", (target_id,))
        conn.commit()

    await update.message.reply_text(f"✅ ইউজার ````{target_id}```` কে আনব্যান করা হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_add_bal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    parts = text.split()
    if len(parts) != 2 or not parts[0].isdigit():
        await update.message.reply_text("❌ সঠিক ফরম্যাটে পাঠাবে (যেমন: `8001997389 50`)")
        return STATE_ADMIN_ADD_BAL

    target_id = int(parts[0])
    try:
        amount = float(parts[1])
    except ValueError:
        await update.message.reply_text("❌ সঠিক টাকার পরিমাণ লিখুন।")
        return STATE_ADMIN_ADD_BAL

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (amount, target_id))
        conn.commit()

    try:
        await context.bot.send_message(
            chat_id=target_id,
            text=f"🎉 <b>আপনার ওয়ালেটে ৳{amount:.2f} এডমিন কর্তৃক যুক্ত করা হয়েছে।</b>",
            parse_mode=ParseMode.HTML
        )
    except Exception:
        pass

    await update.message.reply_text(f"✅ ইউজার ````{target_id}```` এর ওয়ালেটে ৳{amount:.2f} সফলভাবে যোগ করা হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def process_admin_sub_bal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return ConversationHandler.END

    text = update.message.text.strip()
    parts = text.split()
    if len(parts) != 2 or not parts[0].isdigit():
        await update.message.reply_text("❌ সঠিক ফরম্যাটে পাঠাবে (যেমন: `8001997389 20`)")
        return STATE_ADMIN_SUB_BAL

    target_id = int(parts[0])
    try:
        amount = float(parts[1])
    except ValueError:
        await update.message.reply_text("❌ সঠিক টাকার পরিমাণ লিখুন।")
        return STATE_ADMIN_SUB_BAL

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET balance = MAX(0.0, balance - ?) WHERE user_id = ?", (amount, target_id))
        conn.commit()

    try:
        await context.bot.send_message(
            chat_id=target_id,
            text=f"⚠️ <b>আপনার ওয়ালেট থেকে ৳{amount:.2f} কেটে নেওয়া হয়েছে।</b>",
            parse_mode=ParseMode.HTML
        )
    except Exception:
        pass

    await update.message.reply_text(f"✅ ইউজার ````{target_id}```` এর ওয়ালেট থেকে ৳{amount:.2f} কেটে নেওয়া হয়েছে।", reply_markup=get_admin_keyboard(user_id), parse_mode=ParseMode.MARKDOWN)
    clear_user_state(context)
    return ConversationHandler.END

async def cancel_operation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_user_state(context)
    user_id = update.effective_user.id
    await update.message.reply_text(tr(user_id, "operation_cancelled"), reply_markup=get_main_keyboard(user_id))
    return ConversationHandler.END

async def global_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Global error handler to prevent internal error notices to users"""
    logger.error("Exception while handling an update:", exc_info=context.error)

# -----------------------------------------------------------------------------
# MAIN APPLICATION SETUP
# -----------------------------------------------------------------------------
def main():
    logger.info("Initializing bot database and environment...")
    init_db()

    token = "8807027212:AAFSJQnn4llh9w4G9l9aeNzcXT491jDZw5c"
    if not token:
        raise RuntimeError("BOT_TOKEN is not set. Add your NEW BotFather token as the BOT_TOKEN environment variable.")
    app = ApplicationBuilder().token(token).build()

    app.add_error_handler(global_error_handler)

    # Regex filters for bilingual buttons
    btn_task_re = r"^(💼 কাজ \(Task\)|💼 Task)$"
    btn_wallet_re = r"^(👛 ওয়ালেট \(Wallet\)|👛 Wallet)$"
    btn_withdraw_re = r"^(💸 টাকা উত্তোলন \(Withdraw\)|💸 Withdraw)$"
    btn_ref_re = r"^(👥 আমার রেফারেল|👥 My Referrals)$"
    btn_leaderboard_re = r"^(🏆 লিডারবোর্ড|🏆 Leaderboard)$"
    btn_support_re = r"^(🎧 সাপোর্ট|🎧 Support)$"
    btn_lang_re = r"^(🌐 ভাষা পরিবর্তন|🌐 Language)$"
    btn_admin_re = r"^(🟢 Admin Panel|🔑 Admin Panel)$"
    btn_main_menu_re = r"^(🟢 মূল মেনু|🔙 মূল মেনু|🟢 Main Menu|🔙 Main Menu)$"
    btn_cancel_re = r"^(❌ বাতিল|❌ Cancel)$"
    btn_done_re = r"^(✅ জমা দিন \(Done\)|✅ Submit \(Done\)|✅ Done|Done)$"

    btn_ig_re = r"^(📸 ইনস্টাগ্রাম|📸 Instagram)$"
    btn_fb_re = r"^(📘 ফেসবুক|📘 Facebook)$"
    btn_gmail_re = r"^(✉️ জিমেইল|✉️ Gmail)$"

    btn_ig_2fa_re = r"^(📸 ইনস্টাগ্রাম ২এফএ|📸 Instagram 2FA)$"
    btn_ig_seed_re = r"^(📸 ইনস্টাগ্রাম সীড|📸 Instagram Seed)$"
    btn_fb_2fa_re = r"^(📘 ফেসবুক ২এফএ|📘 Facebook 2FA)$"
    btn_fb_cookies_re = r"^(📘 ফেসবুক কুকিজ|📘 Facebook Cookies)$"
    btn_gen_2fa_re = r"^(🔑 2FA Code Generate|🔑 2FA Code Generate করুন|🔑 Generate 2FA Code)$"
    btn_acct_done_re = r"^(✅ একাউন্ট খোলা শেষ|✅ Account Creation Done)$"

    btn_bkash_re = r"^(🟢 বিকাশ \(Bkash\)|🟢 Bkash)$"
    btn_nagad_re = r"^(🟠 নগদ \(Nagad\)|🟠 Nagad)$"
    btn_usdt_re = r"^(🔵 USDT \(BEP-20\))$"

    btn_adm_pending_re = r"^(🟢 📩 পেন্ডিং টাস্ক|📩 পেন্ডিং টাস্ক|🟢 📩 Pending Tasks|📩 Pending Tasks)$"
    btn_adm_held_re = r"^(🟢 ⌛ হোল্ড টাস্ক|⌛ হোল্ড টাস্ক|🟢 ⌛ Held Tasks|⌛ Held Tasks)$"
    btn_adm_search_re = r"^(🟢 🔍 টাস্ক সার্চ|🔍 টাস্ক সার্চ|🟢 🔍 Task Search|🔍 Task Search)$"
    btn_adm_export_re = r"^(🟢 📂 ইউজার এক্সপোর্ট|📂 ইউজার এক্সপোর্ট|🟢 📂 User Export|📂 User Export)$"
    btn_adm_pwd_re = r"^(🟢 🔑 পাসওয়ার্ড ম্যানেজ|🔑 পাসওয়ার্ড ম্যানেজ|🟢 🔑 Password Manager|🔑 Password Manager)$"
    btn_adm_price_re = r"^(🟢 🏷️ প্রাইস ম্যানেজ|🏷️ প্রাইস ম্যানেজ|🟢 🏷️ Price Manager|🏷️ Price Manager)$"
    btn_adm_wd_re = r"^(🟢 💸 উইথড্র রিকোয়েস্ট|💸 উইথড্র রিকোয়েস্ট|🟢 💸 Withdraw Requests|💸 Withdraw Requests)$"
    btn_adm_sys_re = r"^(🟢 ⚙️ সিস্টেম সেটিংস|⚙️ সিস্টেম সেটিংস|🟢 ⚙️ System Settings|⚙️ System Settings)$"
    btn_adm_stats_re = r"^(🟢 📊 লাইভ স্ট্যাটাস|📊 লাইভ স্ট্যাটাস|🟢 📊 Live Stats|📊 Live Stats)$"
    btn_adm_bcast_re = r"^(🟢 📢 ব্রডকাস্ট|📢 ব্রডকাস্ট|🟢 📢 Broadcast|📢 Broadcast)$"

    btn_adm_bulk_app_re = r"^(🟢 ✅ বাল্ক এপ্রুভ|✅ Approve All \(বাল্ক এপ্রুভ\)|🟢 ✅ Approve All|✅ Approve All)$"
    btn_adm_bulk_rej_re = r"^(🟢 ❌ বাল্ক রিজেক্ট|❌ Reject All \(বাল্ক রিজেক্ট\)|🟢 ❌ Reject All|❌ Reject All)$"
    btn_adm_excel_re = r"^(🟢 📊 এক্সেল শীট \(Spreadsheet\)|📊 এক্সেল শীট \(Spreadsheet\)|🟢 📊 Spreadsheet UI|📊 Spreadsheet UI)$"
    btn_adm_user_mgmt_re = r"^(🟢 👥 ইউজার ম্যানেজমেন্ট|👥 ইউজার ম্যানেজমেন্ট|🟢 👥 User Management|👥 User Management)$"
    btn_adm_admin_mgmt_re = r"^(🟢 👑 এডমিন কন্ট্রোল|👑 এডমিন কন্ট্রোল|🟢 👑 Admin Control|👑 Admin Control)$"

    # --- Main Navigation Handlers ---
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(MessageHandler(filters.Regex(btn_task_re), handle_task_menu))
    app.add_handler(MessageHandler(filters.Regex(btn_wallet_re), handle_wallet_menu))
    app.add_handler(MessageHandler(filters.Regex(btn_ref_re), handle_referrals))
    app.add_handler(MessageHandler(filters.Regex(btn_leaderboard_re), handle_leaderboard))
    app.add_handler(MessageHandler(filters.Regex(btn_support_re), handle_support))
    app.add_handler(MessageHandler(filters.Regex(btn_lang_re), handle_language))
    app.add_handler(MessageHandler(filters.Regex(btn_admin_re), handle_admin_panel))
    app.add_handler(MessageHandler(filters.Regex(btn_main_menu_re), start_command))

    # --- Admin ReplyKeyboard Buttons ---
    app.add_handler(MessageHandler(filters.Regex(btn_adm_pending_re), admin_pending_tasks))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_held_re), admin_held_tasks))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_pwd_re), admin_password_manager))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_price_re), admin_price_manager))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_wd_re), admin_withdraw_requests))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_sys_re), admin_system_settings))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_stats_re), admin_live_stats))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_excel_re), admin_excel_export_menu))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_user_mgmt_re), admin_user_management))
    app.add_handler(MessageHandler(filters.Regex(btn_adm_admin_mgmt_re), admin_control_panel))

    # --- Task Conversation Handler ---
    task_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex(btn_ig_re), handle_instagram_menu),
            MessageHandler(filters.Regex(btn_fb_re), handle_facebook_menu),
            MessageHandler(filters.Regex(btn_gmail_re), handle_gmail_menu),
        ],
        states={
            STATE_TASK_CATEGORY: [
                MessageHandler(filters.Regex(btn_ig_re), handle_instagram_menu),
                MessageHandler(filters.Regex(btn_fb_re), handle_facebook_menu),
                MessageHandler(filters.Regex(btn_gmail_re), handle_gmail_menu),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
            ],
            STATE_TASK_INSTAGRAM: [
                MessageHandler(filters.Regex(btn_ig_2fa_re), handle_ig_2fa_task),
                MessageHandler(filters.Regex(btn_ig_seed_re), handle_ig_seed_task),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
            ],
            STATE_TASK_FACEBOOK: [
                MessageHandler(filters.Regex(btn_fb_2fa_re), handle_fb_2fa_task),
                MessageHandler(filters.Regex(btn_fb_cookies_re), handle_fb_cookies_task),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
            ],
            STATE_TASK_IG_2FA_INPUT: [
                MessageHandler(filters.Regex(btn_gen_2fa_re), process_2fa_secret),
                MessageHandler(filters.Regex(btn_acct_done_re), process_task_finalize),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_2fa_secret),
            ],
            STATE_TASK_FB_2FA_INPUT: [
                MessageHandler(filters.Regex(btn_gen_2fa_re), process_2fa_secret),
                MessageHandler(filters.Regex(btn_acct_done_re), process_task_finalize),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_2fa_secret),
            ],
            STATE_TASK_2FA_SECRET: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_2fa_secret),
            ],
            STATE_TASK_FINAL_SUBMIT: [
                MessageHandler(filters.Regex(btn_gen_2fa_re), process_2fa_secret),
                MessageHandler(filters.Regex(btn_acct_done_re), process_task_finalize),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_task_finalize),
            ],
            STATE_TASK_IG_SEED_INPUT: [
                MessageHandler(filters.Regex(btn_done_re), process_ig_seed_finalize),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.Document.ALL, handle_ig_seed_file_or_text),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_ig_seed_file_or_text),
            ],
            STATE_TASK_FB_COOKIES_UID: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_fb_cookies_uid),
            ],
            STATE_TASK_FB_COOKIES_STR: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_fb_cookies_str),
            ],
            STATE_TASK_GMAIL_INPUT: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_gmail_submission),
            ],
        },
        fallbacks=[
            CommandHandler("start", start_command),
            MessageHandler(filters.Regex(btn_main_menu_re), start_command),
        ],
        per_message=False,
        allow_reentry=True
    )
    app.add_handler(task_conv)

    # --- Withdraw Conversation Handler ---
    withdraw_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex(btn_withdraw_re), handle_withdraw_start),
        ],
        states={
            STATE_WD_METHOD: [
                MessageHandler(filters.Regex(btn_bkash_re), handle_withdraw_method),
                MessageHandler(filters.Regex(btn_nagad_re), handle_withdraw_method),
                MessageHandler(filters.Regex(btn_usdt_re), handle_withdraw_method),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
            ],
            STATE_WD_NUMBER: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_number),
            ],
            STATE_WD_AMOUNT: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_amount),
            ],
        },
        fallbacks=[
            CommandHandler("start", start_command),
            MessageHandler(filters.Regex(btn_main_menu_re), start_command),
            MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
        ],
        per_message=False,
        allow_reentry=True
    )
    app.add_handler(withdraw_conv)

    # --- Admin Input Conversation Handler ---
    admin_text_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex(btn_adm_search_re), admin_search_prompt),
            MessageHandler(filters.Regex(btn_adm_export_re), admin_export_prompt),
            MessageHandler(filters.Regex(btn_adm_bcast_re), admin_broadcast_prompt),
            MessageHandler(filters.Regex(btn_adm_bulk_app_re), admin_bulk_approve_prompt),
            MessageHandler(filters.Regex(btn_adm_bulk_rej_re), admin_bulk_reject_prompt),
            CallbackQueryHandler(admin_callback_handler, pattern="^(adm_usr_ban|adm_usr_unban|adm_usr_add_bal|adm_usr_sub_bal|adm_usr_search|adm_add_subadmin|adm_rem_subadmin|adm_edit_|adm_edit_force_channels_list)"),
        ],
        states={
            STATE_ADMIN_SEARCH_QUERY: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_text),
            ],
            STATE_ADMIN_EXPORT_USER: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_text),
            ],
            STATE_ADMIN_BROADCAST: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.ALL & ~filters.COMMAND, process_admin_broadcast_media),
            ],
            STATE_ADMIN_REJECT_REASON: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_text),
            ],
            STATE_ADMIN_SETTING_VAL: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_text),
            ],
            STATE_ADMIN_BULK_APPROVE: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_bulk_approve),
            ],
            STATE_ADMIN_BULK_REJECT: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_bulk_reject),
            ],
            STATE_ADMIN_ADD_SUBADMIN: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_add_subadmin),
            ],
            STATE_ADMIN_REM_SUBADMIN: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_rem_subadmin),
            ],
            STATE_ADMIN_BAN_USER: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_ban_user),
            ],
            STATE_ADMIN_UNBAN_USER: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_unban_user),
            ],
            STATE_ADMIN_ADD_BAL: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_add_bal),
            ],
            STATE_ADMIN_SUB_BAL: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_sub_bal),
            ],
            STATE_ADMIN_USER_SEARCH: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_user_search),
            ],
            STATE_ADMIN_FORCE_CHANNELS_EDIT: [
                MessageHandler(filters.Regex(btn_cancel_re), cancel_operation),
                MessageHandler(filters.Regex(btn_main_menu_re), start_command),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_admin_force_channels_edit),
            ],
        },
        fallbacks=[
            CommandHandler("start", start_command),
            MessageHandler(filters.Regex(btn_main_menu_re), start_command),
        ],
        per_message=False,
        allow_reentry=True
    )
    app.add_handler(admin_text_conv)

    # --- Global Callback Query Handlers ---
    app.add_handler(CallbackQueryHandler(language_callback, pattern="^(lang_bn|lang_en)$"))
    app.add_handler(CallbackQueryHandler(check_force_join_cb, pattern="^check_force_join_cb$"))
    app.add_handler(CallbackQueryHandler(admin_excel_export_download_cb, pattern="^adm_xlsx_"))
    app.add_handler(CallbackQueryHandler(admin_excel_clear_cb, pattern="^adm_cls_"))
    app.add_handler(CallbackQueryHandler(
        admin_callback_handler,
        pattern="^(adm_app_sub_|adm_hold_sub_|adm_rej_prompt_|adm_rej_act_|adm_rej_custom_|adm_app_wd_|adm_rej_wd_|adm_back_to_admin|adm_toggle_|adm_edit_|adm_add_subadmin|adm_rem_subadmin|adm_usr_)"
    ))

    logger.info("Trust Vault Mails Style Bot successfully started!")
    app.run_polling()

if __name__ == "__main__":
    main()
